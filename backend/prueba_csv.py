import io
import csv
import re
import json
import datetime
import calendar
from pathlib import Path
import openpyxl
from googleapiclient.http import MediaIoBaseDownload
from dateparser import parse
from config.config import drive

MESES_MAP = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4,
    "MAYO": 5, "JUNIO": 6, "JULIO": 7, "AGOSTO": 8,
    "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12
}

def parse_date_from_filename(name):
    pattern = r"(\d{1,2})\s+(?:de\s+)?(ENERO|FEBRERO|MARZO|ABRIL|MAYO|JUNIO|JULIO|AGOSTO|SEPTIEMBRE|OCTUBRE|NOVIEMBRE|DICIEMBRE)\s+(\d{4})"
    match = re.search(pattern, name, re.IGNORECASE)
    if match:
        day = int(match.group(1))
        month_str = match.group(2).upper()
        year = int(match.group(3))
        month = MESES_MAP.get(month_str, 1)
        try:
            return datetime.date(year, month, day).isoformat()
        except ValueError:
            pass
            
    pattern_no_year = r"(\d{1,2})\s+(?:de\s+)?(ENERO|FEBRERO|MARZO|ABRIL|MAYO|JUNIO|JULIO|AGOSTO|SEPTIEMBRE|OCTUBRE|NOVIEMBRE|DICIEMBRE)"
    match_ny = re.search(pattern_no_year, name, re.IGNORECASE)
    if match_ny:
        day = int(match_ny.group(1))
        month_str = match_ny.group(2).upper()
        month = MESES_MAP.get(month_str, 1)
        try:
            return datetime.date(2026, month, day).isoformat()
        except ValueError:
            pass
            
    return None

def excel_to_csv_in_memory(file_item):
    # 1. Descargar Excel de Drive
    request = drive.files().get_media(fileId=file_item['id'])
    excel_stream = io.BytesIO()
    downloader = MediaIoBaseDownload(excel_stream, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    excel_stream.seek(0)
    
    # 2. Leer con openpyxl (modo lectura rápida)
    wb = openpyxl.load_workbook(excel_stream, read_only=True, data_only=True)
    if "DEMOSTRATIVO" not in wb.sheetnames:
        raise ValueError("No se encontró la hoja 'DEMOSTRATIVO' en el Excel.")
    sheet = wb["DEMOSTRATIVO"]
    
    # 3. Escribir a CSV en memoria (string stream)
    csv_stream = io.StringIO()
    writer = csv.writer(csv_stream)
    for row in sheet.iter_rows(values_only=True):
        if any(v is not None for v in row):
            writer.writerow(row)
            
    csv_stream.seek(0)
    return csv_stream.getvalue()

def extraer_datos_csv(csv_content):
    f_stream = io.StringIO(csv_content)
    reader = csv.reader(f_stream)
    rows = list(reader)
    
    CAMPOS = {"APELLIDOS Y NOMBRES", "CEDULA", "SUBNOVEDAD", "DESCRIPCION", "DESDE", "HASTA"}
    headers_map = {}
    fila_encabezado_idx = -1
    
    # Buscar cabeceras
    for idx, row in enumerate(rows):
        if idx > 50:
            break
        encabezados = {}
        for col_idx, val in enumerate(row):
            if val:
                val_clean = val.strip().upper()
                if val_clean in CAMPOS:
                    encabezados[val_clean] = col_idx
        if "CEDULA" in encabezados:
            headers_map = encabezados
            fila_encabezado_idx = idx
            break
            
    if fila_encabezado_idx == -1:
        raise ValueError("No se encontró la fila de encabezados en el CSV.")
        
    datos = {}
    for idx in range(fila_encabezado_idx + 1, len(rows)):
        row = rows[idx]
        if not row or len(row) <= max(headers_map.values()):
            continue
            
        cedula_val = row[headers_map["CEDULA"]]
        nombre_val = row[headers_map["APELLIDOS Y NOMBRES"]]
        subnovedad_val = row[headers_map["SUBNOVEDAD"]]
        
        if not cedula_val or not nombre_val:
            continue
            
        try:
            cedula_int = int(float(str(cedula_val).strip()))
        except ValueError:
            continue
            
        if cedula_int <= 0:
            continue
            
        descripcion = ""
        if "DESCRIPCION" in headers_map:
            descripcion = str(row[headers_map["DESCRIPCION"]]).strip() if row[headers_map["DESCRIPCION"]] else ""
            
        desde = ""
        if "DESDE" in headers_map:
            desde = str(row[headers_map["DESDE"]]).strip() if row[headers_map["DESDE"]] else ""
            
        hasta = ""
        if "HASTA" in headers_map:
            hasta = str(row[headers_map["HASTA"]]).strip() if row[headers_map["HASTA"]] else ""
            
        datos[str(cedula_int)] = {
            "CEDULA": cedula_int,
            "APELLIDOS Y NOMBRES": str(nombre_val).strip().upper(),
            "SUBNOVEDAD": str(subnovedad_val).strip().upper() if subnovedad_val else "SIN NOVEDAD",
            "DESCRIPCION": descripcion,
            "DESDE": desde,
            "HASTA": hasta
        }
    return datos

def obtener_hojas_csv_test(target_month=None, target_date=None, force_overwrite=False):
    import time
    from leer_archivos_excel import consultar_fechas_db
    
    t_global_start = time.time()
    
    with open("listado_meses.json", encoding="utf-8") as l:
        listado_meses = json.load(l)

    errors = []
    dates_in_db = consultar_fechas_db()

    months_to_process = []
    if target_month:
        months_to_process = [target_month]
    elif target_date:
        try:
            m_num = int(target_date.split("-")[1])
            reverse_map = {
                1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
                5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
                9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
            }
            months_to_process = [reverse_map.get(m_num)]
        except Exception:
            months_to_process = list(listado_meses.keys())
    else:
        months_to_process = list(listado_meses.keys())

    for mes in months_to_process:
        if mes not in MESES_MAP:
            continue
            
        t_month_start = time.time()
        archivoF = Path(f"listadoMeses/{mes}.json")
        archivos = listado_meses.get(mes, [])
        
        datos_archivo = {}
        if archivoF.exists():
            try:
                with open(archivoF, "r", encoding="utf-8") as f:
                    datos_archivo = json.load(f)
                print(f"[CSV Test] El mes {mes} ya tiene {len(datos_archivo)} días cargados localmente.")
            except Exception as e:
                datos_archivo = {}

        month_num = MESES_MAP[mes]
        num_days = calendar.monthrange(2026, month_num)[1]
        all_month_dates = [datetime.date(2026, month_num, d).isoformat() for d in range(1, num_days + 1)]
        
        if target_date:
            if force_overwrite:
                missing_dates = [target_date]
            else:
                missing_dates = [target_date] if (target_date not in dates_in_db and target_date not in datos_archivo) else []
        else:
            if force_overwrite:
                missing_dates = all_month_dates
            else:
                missing_dates = [d for d in all_month_dates if d not in dates_in_db and d not in datos_archivo]
        
        if not missing_dates:
            print(f"[CSV Test] El mes/día de {mes} está completamente cargado. Omitiendo.")
            continue
            
        drive_files_by_date = {}
        for archivo in archivos:
            fecha_parsed = parse_date_from_filename(archivo['name'])
            if fecha_parsed:
                drive_files_by_date[fecha_parsed] = archivo
                
        files_to_download = []
        for m_date in missing_dates:
            if m_date in drive_files_by_date:
                files_to_download.append((m_date, drive_files_by_date[m_date]))
                
        nuevos_datos_cargados = False
        for f_date, archivo in files_to_download:
            print(f"[CSV Test] Descargando y convirtiendo {f_date}: {archivo['name']}")
            t_down_start = time.time()
            try:
                # Conversión en memoria a CSV
                csv_content = excel_to_csv_in_memory(archivo)
                t_conv = time.time()
                print(f"  [CSV Test] Descarga y conversión a CSV completada en {t_conv - t_down_start:.2f} segundos.")
                
                # Lectura desde CSV
                t_parse_start = time.time()
                datos_archivo[f_date] = extraer_datos_csv(csv_content)
                t_parse_end = time.time()
                print(f"  [CSV Test] Parseo de CSV en memoria completado en {t_parse_end - t_parse_start:.4f} segundos.")
                nuevos_datos_cargados = True
            except Exception as e:
                print(f"[CSV Test] Error procesando {archivo['name']}: {e}")
                errors.append({"file": archivo['name'], "error": f"Error CSV Test: {str(e)}"})

        if nuevos_datos_cargados or not archivoF.exists():
            datos_archivo = dict(sorted(datos_archivo.items()))
            with open(f"listadoMeses/{mes}.json", "w", encoding="utf-8") as d:
                json.dump(datos_archivo, d, indent=4, ensure_ascii=False, default=str)
            
        t_month_end = time.time()
        print(f"[CSV Test] Sincronización del mes {mes} terminada. Tiempo: {t_month_end - t_month_start:.2f} segundos.")
            
    t_global_end = time.time()
    print(f"[CSV Test Global] Tiempo total acumulado: {t_global_end - t_global_start:.2f} segundos.")
    return errors
