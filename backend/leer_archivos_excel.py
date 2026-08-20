from openpyxl import load_workbook
import os
import io
import json
from googleapiclient.http import MediaIoBaseDownload


from config.auth import obtener_servicio_drive
import openpyxl
import re
from dateparser import parse
from pathlib import Path
import datetime
import calendar
import csv

# Mapa para convertir el nombre del mes en su número correlativo
MESES_MAP = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4,
    "MAYO": 5, "JUNIO": 6, "JULIO": 7, "AGOSTO": 8,
    "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12
}

def parse_date_from_filename(name):
    """
    Intenta extraer la fecha en formato YYYY-MM-DD a partir del nombre del archivo en Drive.
    Soporta:
    - '01 DE AGOSTO 2026', '1 AGOSTO 2026', '1 AGOSTO'
    - '2026-08-01', '2026_08_01'
    - '01-08-2026', '01_08_2026', '01.08.2026'
    """
    # 1. Formato con texto de mes y año: '01 de Agosto 2026' o '1 Agosto 2026'
    pattern = r"(\d{1,2})\s+(?:de\s+)?(ENERO|FEBRERO|MARZO|ABRIL|MAYO|JUNIO|JULIO|AGOSTO|SEPTIEMBRE|OCTUBRE|NOVIEMBRE|DICIEMBRE)\s+(\d{4})"
    match = re.search(pattern, name, re.IGNORECASE)
    if match:
        day = int(match.group(1))
        month_str = match.group(2).upper()
        year = int(match.group(3))
        month = MESES_MAP.get(month_str, 1)
        try:
            return datetime.date(year, month, day).isoformat()
        except ValueError:
            pass

    # 2. Formato con texto de mes sin año: '01 de Agosto' o '1 Agosto' (asume año 2026)
    pattern_no_year = r"(\d{1,2})\s+(?:de\s+)?(ENERO|FEBRERO|MARZO|ABRIL|MAYO|JUNIO|JULIO|AGOSTO|SEPTIEMBRE|OCTUBRE|NOVIEMBRE|DICIEMBRE)"
    match_ny = re.search(pattern_no_year, name, re.IGNORECASE)
    if match_ny:
        day = int(match_ny.group(1))
        month_str = match_ny.group(2).upper()
        month = MESES_MAP.get(month_str, 1)
        try:
            return datetime.date(2026, month, day).isoformat()
        except ValueError:
            pass

    # 3. Formato numérico ISO: '2026-08-01' o '2026_08_01'
    iso_pattern = r"(202\d)[-_](\d{1,2})[-_](\d{1,2})"
    match_iso = re.search(iso_pattern, name)
    if match_iso:
        y, m, d = int(match_iso.group(1)), int(match_iso.group(2)), int(match_iso.group(3))
        try:
            return datetime.date(y, m, d).isoformat()
        except ValueError:
            pass

    # 4. Formato numérico latino: '01-08-2026' o '01_08_2026' o '01.08.2026'
    lat_pattern = r"(\d{1,2})[-_./](\d{1,2})[-_./](202\d)"
    match_lat = re.search(lat_pattern, name)
    if match_lat:
        d, m, y = int(match_lat.group(1)), int(match_lat.group(2)), int(match_lat.group(3))
        try:
            return datetime.date(y, m, d).isoformat()
        except ValueError:
            pass

    return None

def excel_to_csv_in_memory(file_item, drive=None):
    """
    Descarga el archivo Excel desde Google Drive en memoria RAM,
    lee la hoja 'DEMOSTRATIVO' (o la primera hoja disponible) usando openpyxl
    y convierte el contenido a un string con formato CSV estándar.
    """
    import time
    if drive is None:
        drive = obtener_servicio_drive()
    
    # 1. Petición de descarga de binarios de Drive
    t_down_start = time.time()
    request = drive.files().get_media(fileId=file_item['id'])

    excel_stream = io.BytesIO()
    downloader = MediaIoBaseDownload(excel_stream, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    excel_stream.seek(0)
    t_down_end = time.time()
    print(f"  [Timer Detail] Descarga de Google Drive: {t_down_end - t_down_start:.2f} segundos.")
    
    # 2. Cargar Excel en modo rápido (read_only=True y data_only=True)
    t_open_start = time.time()
    wb = openpyxl.load_workbook(excel_stream, read_only=True, data_only=True)
    
    target_sheet = None
    for s_name in wb.sheetnames:
        if "demostrativo" in s_name.strip().lower():
            target_sheet = s_name
            break
            
    if not target_sheet and len(wb.sheetnames) > 0:
        target_sheet = wb.sheetnames[0]
        
    if not target_sheet:
        raise ValueError("No se encontraron hojas de cálculo en el libro de Excel.")
        
    sheet = wb[target_sheet]
    t_open_end = time.time()
    print(f"  [Timer Detail] Apertura de Excel '{target_sheet}' (openpyxl): {t_open_end - t_open_start:.2f} segundos.")
    
    # 3. Escribir a una cadena de texto CSV en memoria RAM
    t_write_start = time.time()
    csv_stream = io.StringIO()
    writer = csv.writer(csv_stream)
    for row in sheet.iter_rows(values_only=True):
        # Omitir filas que estén completamente vacías
        if any(v is not None for v in row):
            writer.writerow(row)
            
    csv_stream.seek(0)
    t_write_end = time.time()
    print(f"  [Timer Detail] Conversión y Escritura a CSV: {t_write_end - t_write_start:.4f} segundos.")
    
    return csv_stream.getvalue()


def extraer_datos_csv(csv_content):
    """
    Parsea los datos del string CSV en memoria usando el módulo nativo 'csv' de Python.
    La lectura de CSV en Python está escrita en C, por lo que iterar las filas toma milisegundos.
    
    Diferencias en lectura:
    - En lugar de acceder a coordenadas bidimensionales de openpyxl (lento),
      accedemos por indexación directa de arreglos (lista de strings: `row[col_idx]`), lo cual es instantáneo.
    """
    f_stream = io.StringIO(csv_content)
    reader = csv.reader(f_stream)
    rows = list(reader)
    
    CAMPOS = {"APELLIDOS Y NOMBRES", "CEDULA", "SUBNOVEDAD", "DESCRIPCION", "DESDE", "HASTA"}
    headers_map = {}
    fila_encabezado_idx = -1
    
    # Búsqueda dinámica de la fila de cabeceras (dentro de las primeras 50 líneas)
    for idx, row in enumerate(rows):
        if idx > 50:
            break
        encabezados = {}
        for col_idx, val in enumerate(row):
            if val:
                val_clean = val.strip().upper()
                if val_clean in CAMPOS:
                    encabezados[val_clean] = col_idx
        if "CEDULA" in encabezados:
            headers_map = encabezados
            fila_encabezado_idx = idx
            break
            
    if fila_encabezado_idx == -1:
        raise ValueError("No se encontró la fila de encabezados en el CSV.")
        
    datos = {}
    # Recorrer los registros de personal
    for idx in range(fila_encabezado_idx + 1, len(rows)):
        row = rows[idx]
        if not row or len(row) <= max(headers_map.values()):
            continue
            
        cedula_val = row[headers_map["CEDULA"]]
        nombre_val = row[headers_map["APELLIDOS Y NOMBRES"]]
        subnovedad_val = row[headers_map["SUBNOVEDAD"]]
        
        if not cedula_val or not nombre_val:
            continue
            
        try:
            cedula_int = int(float(str(cedula_val).strip()))
        except ValueError:
            continue
            
        if cedula_int <= 0:
            continue
            
        descripcion = ""
        if "DESCRIPCION" in headers_map:
            descripcion = str(row[headers_map["DESCRIPCION"]]).strip() if row[headers_map["DESCRIPCION"]] else ""
            
        desde = ""
        if "DESDE" in headers_map:
            desde = str(row[headers_map["DESDE"]]).strip() if row[headers_map["DESDE"]] else ""
            
        hasta = ""
        if "HASTA" in headers_map:
            hasta = str(row[headers_map["HASTA"]]).strip() if row[headers_map["HASTA"]] else ""
            
        datos[str(cedula_int)] = {
            "CEDULA": cedula_int,
            "APELLIDOS Y NOMBRES": str(nombre_val).strip().upper(),
            "SUBNOVEDAD": str(subnovedad_val).strip().upper() if subnovedad_val else "SIN NOVEDAD",
            "DESCRIPCION": descripcion,
            "DESDE": desde,
            "HASTA": hasta
        }
    return datos

def consultar_fechas_db():
    """
    Abre una conexión temporal rápida a base de datos para traer las fechas ya sincronizadas,
    y la cierra de inmediato para evitar que quede inactiva.
    """
    dates_in_db = set()
    try:
        from app.database import ConnectionWrapper, DB_CONN_PARAMS
        print(f"Consultando fechas existentes en PostgreSQL local ({DB_CONN_PARAMS.get('dbname')})...")
        temp_conn = ConnectionWrapper(conn_params=DB_CONN_PARAMS)
        temp_cursor = temp_conn.cursor()
        temp_cursor.execute("SELECT fecha FROM REPORTES;")
        dates_in_db = {row[0] for row in temp_cursor.fetchall()}
        temp_conn.close()
        print(f"Fechas obtenidas con éxito: {len(dates_in_db)} encontradas.")
    except Exception as e:
        print(f"Error consultando fechas en base de datos: {e}")
        dates_in_db = set()
    return dates_in_db

def descargar_y_procesar_fecha(archivo, f_date, datos_archivo, drive=None):
    """
    Descarga el archivo Excel de Drive, lo convierte a CSV en memoria y estructura
    los datos dentro de 'datos_archivo'. Registra métricas de tiempo por consola.
    """
    import time
    print(f"Descargando y procesando {f_date}: {archivo['name']}")
    t_down_start = time.time()
    try:
        # Conversión en memoria a CSV
        csv_content = excel_to_csv_in_memory(archivo, drive=drive)
        t_conv = time.time()
        print(f"  [Timer] Descarga y conversión a CSV completada en {t_conv - t_down_start:.2f} segundos.")
        
        # Lectura desde CSV
        t_parse_start = time.time()
        datos_archivo[f_date] = extraer_datos_csv(csv_content)
        t_parse_end = time.time()
        print(f"  [Timer] Parseo de CSV en memoria completado en {t_parse_end - t_parse_start:.4f} segundos.")
        return True
    except Exception as e:
        print(f"Error procesando {archivo['name']}: {e}")
        raise e

def descargar_y_procesar_fallback(archivo, still_missing, datos_archivo, drive=None):
    """
    Procesa archivos que no tienen fecha en su nombre. Descarga, convierte a CSV en memoria,
    inspecciona su contenido buscando la fecha y procesa si coincide con alguna fecha faltante.
    Retorna la fecha encontrada (str) o None si no se encuentra o es inválida.
    """
    import time
    t_fallback_start = time.time()
    try:
        # Descarga y conversión a CSV en memoria
        csv_content = excel_to_csv_in_memory(archivo, drive=drive)
    except Exception as e:
        print(f"Error descargando archivo fallback {archivo['name']}: {e}")
        raise e

        
    f_stream = io.StringIO(csv_content)
    reader = csv.reader(f_stream)
    rows = list(reader)
    
    # Extraer la fecha buscando la celda que contiene 'venecia - caqueta'
    fecha_archivo = ""
    for row in rows:
        for val in row:
            if val and "venecia - caqueta" in val.lower():
                texto = val
                m = re.search(r"\d.*", texto)
                if m:
                    fecha_archivo = re.sub(r'(\d{1,2}),\s+de', r'\1 de', m.group())
                    break
        if fecha_archivo:
            break
            
    if not fecha_archivo:
        return None
        
    try:
        fecha = parse(fecha_archivo).date().isoformat()
    except Exception:
        return None
        
    if fecha in still_missing:
        try:
            datos_archivo[fecha] = extraer_datos_csv(csv_content)
            t_fallback_end = time.time()
            print(f"  -> Día {fecha} cargado exitosamente (fallback) en {t_fallback_end - t_fallback_start:.2f} segundos.")
            still_missing.remove(fecha)
            return fecha
        except Exception as e:
            print(f"Error al extraer estructura en fallback para {archivo['name']}: {e}")
            raise e
    return fecha

def obtener_hojas(db=None, target_month=None, target_date=None, target_dates=None, force_overwrite=False):
    """
    Función orquestadora principal que gestiona el flujo de descargas de reportes desde Google Drive.
    Filtra los meses/días solicitados, verifica fechas ausentes en la base de datos de PostgreSQL local y
    escribe/actualiza los archivos JSON de listado local en caché.
    """
    import time
    t_global_start = time.time()
    drive = obtener_servicio_drive()

    
    # Carga el mapeo de archivos de Drive guardado por leer_carpetas
    with open("listado_meses.json", encoding="utf-8") as l:
        listado_meses = json.load(l)

    errors = []
    sync_log = [] # Registro amigable para el usuario final sin detalles técnicos
    
    # Obtener fechas cargadas en la base de datos local
    dates_in_db = consultar_fechas_db()

    # Mapa inverso de número de mes a nombre
    reverse_map = {
        1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
        5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
        9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
    }

    # Identificar meses a procesar según filtros
    months_to_process = []
    if target_dates:
        # Modo multi-día: agrupar las fechas seleccionadas por mes
        meses_set = set()
        for td in target_dates:
            try:
                m_num = int(td.split("-")[1])
                mes_name = reverse_map.get(m_num)
                if mes_name:
                    meses_set.add(mes_name)
            except Exception:
                pass
        months_to_process = list(meses_set)
    elif target_month:
        months_to_process = [target_month]
    elif target_date:
        try:
            m_num = int(target_date.split("-")[1])
            months_to_process = [reverse_map.get(m_num)]
        except Exception:
            months_to_process = list(listado_meses.keys())
    else:
        months_to_process = list(listado_meses.keys())

    # Procesar cada mes
    for mes in months_to_process:
        if mes not in MESES_MAP:
            continue
            
        t_month_start = time.time()
        os.makedirs("listadoMeses", exist_ok=True)
        archivoF = Path(f"listadoMeses/{mes}.json")

        archivos = listado_meses.get(mes, [])
        
        datos_archivo = {}
        if archivoF.exists():
            try:
                with open(archivoF, "r", encoding="utf-8") as f:
                    datos_archivo = json.load(f)
                print(f"El mes {mes} ya tiene {len(datos_archivo)} días cargados localmente.")
            except Exception as e:
                datos_archivo = {}

        # Cargar metadatos del mes (caché de nombres de archivos -> fechas procesadas)
        metadataF = Path(f"listadoMeses/{mes}.metadata")
        metadata = {"processed_files": {}}
        if not force_overwrite and metadataF.exists():
            try:
                with open(metadataF, "r", encoding="utf-8") as mf:
                    metadata = json.load(mf)
            except Exception:
                pass
        processed_files = metadata.setdefault("processed_files", {})

        # Fechas esperadas del mes
        month_num = MESES_MAP[mes]
        num_days = calendar.monthrange(2026, month_num)[1]
        all_month_dates = [datetime.date(2026, month_num, d).isoformat() for d in range(1, num_days + 1)]
        
        # Filtrar fechas faltantes
        if target_dates:
            # Modo multi-día: solo las fechas seleccionadas que pertenecen a este mes
            dates_this_month = [d for d in target_dates if d in all_month_dates]
            if force_overwrite:
                missing_dates = dates_this_month
            else:
                missing_dates = [d for d in dates_this_month if d not in dates_in_db and d not in datos_archivo]
        elif target_date:
            if force_overwrite:
                missing_dates = [target_date]
            else:
                missing_dates = [target_date] if (target_date not in dates_in_db and target_date not in datos_archivo) else []
        else:
            if force_overwrite:
                missing_dates = all_month_dates
            else:
                missing_dates = [d for d in all_month_dates if d not in dates_in_db and d not in datos_archivo]
        
        # Registrar archivos que omitimos porque ya se encuentran sincronizados
        # En modo multi-día, solo informar sobre archivos de las fechas seleccionadas por el usuario
        for archivo in archivos:
            fecha_parsed = parse_date_from_filename(archivo['name'])
            if fecha_parsed and fecha_parsed not in missing_dates:
                # Si estamos en modo multi-día, solo loguear si la fecha es una de las seleccionadas
                if target_dates and fecha_parsed not in target_dates:
                    continue
                sync_log.append({
                    "file": archivo['name'],
                    "status": "skipped",
                    "detail": "Omitido: este día ya se encuentra cargado y al día."
                })
        
        if not missing_dates:
            print(f"El mes/día de {mes} está completamente cargado. Omitiendo.")
            # Guardamos los metadatos por si acaso hubo cambios
            try:
                with open(metadataF, "w", encoding="utf-8") as mf:
                    json.dump(metadata, mf, indent=4, ensure_ascii=False)
            except Exception:
                pass
            continue
            
        print(f"Fechas faltantes/requeridas para {mes}: {missing_dates}")

        # Indexar archivos de Drive del mes
        drive_files_by_date = {}
        unparsed_files = []
        
        for archivo in archivos:
            fecha_parsed = parse_date_from_filename(archivo['name'])
            if fecha_parsed:
                drive_files_by_date[fecha_parsed] = archivo
                processed_files[archivo['name']] = fecha_parsed
            else:
                # Es un archivo sin fecha en el nombre (fallback). ¿Ya lo evaluamos antes?
                cached_date = processed_files.get(archivo['name'])
                if cached_date:
                    if cached_date != "unknown":
                        drive_files_by_date[cached_date] = archivo
                    # Si es "unknown" no se hace nada, evitando volver a descargarlo
                else:
                    unparsed_files.append(archivo)
                
        # Emparejar descargas necesarias
        files_to_download = []
        for m_date in missing_dates:
            if m_date in drive_files_by_date:
                files_to_download.append((m_date, drive_files_by_date[m_date]))
                
        # Iniciar descargas y procesamiento
        nuevos_datos_cargados = False
        for f_date, archivo in files_to_download:
            try:
                success = descargar_y_procesar_fecha(archivo, f_date, datos_archivo, drive=drive)
                if success:
                    nuevos_datos_cargados = True
                    sync_log.append({
                        "file": archivo['name'],
                        "status": "success",
                        "detail": f"Sincronizado con éxito como reporte del día {f_date}."
                    })
            except Exception as e:
                errors.append({"file": archivo['name'], "error": f"Error descargando/procesando: {str(e)}"})
                sync_log.append({
                    "file": archivo['name'],
                    "status": "error",
                    "detail": f"Error al procesar: {str(e)}"
                })
                
        # Búsqueda fallback
        still_missing = [d for d in missing_dates if d not in datos_archivo]
        if still_missing and unparsed_files:
            print(f"Hay {len(still_missing)} fechas faltantes y {len(unparsed_files)} archivos sin fecha legible en el nombre. Evaluando fallbacks...")
            for archivo in unparsed_files:
                try:
                    fecha_detectada = descargar_y_procesar_fallback(archivo, still_missing, datos_archivo, drive=drive)
                    if fecha_detectada:
                        processed_files[archivo['name']] = fecha_detectada
                        if fecha_detectada in datos_archivo:
                            nuevos_datos_cargados = True
                            sync_log.append({
                                "file": archivo['name'],
                                "status": "success",
                                "detail": f"Sincronizado con éxito (identificado como reporte de {fecha_detectada})."
                            })
                        else:
                            sync_log.append({
                                "file": archivo['name'],
                                "status": "skipped",
                                "detail": f"Omitido: identificado como reporte de {fecha_detectada} (ya al día)."
                            })
                    else:
                        processed_files[archivo['name']] = "unknown"
                        sync_log.append({
                            "file": archivo['name'],
                            "status": "skipped",
                            "detail": "Omitido: el archivo no corresponde a ninguna fecha faltante."
                        })
                except Exception as e:
                    errors.append({"file": archivo['name'], "error": f"Error en fallback: {str(e)}"})
                    processed_files[archivo['name']] = "unknown"
                    sync_log.append({
                        "file": archivo['name'],
                        "status": "error",
                        "detail": f"Error en lectura de reporte ({str(e)})."
                    })


        # Guardar en local JSON si hay cambios
        if nuevos_datos_cargados or not archivoF.exists():
            datos_archivo = dict(sorted(datos_archivo.items()))
            print(f"Guardando actualizaciones para el mes {mes}...")
            with open(f"listadoMeses/{mes}.json", "w", encoding="utf-8") as d:
                json.dump(datos_archivo, d, indent=4, ensure_ascii=False, default=str)
        else:
            print(f"No se encontraron nuevos reportes para el mes {mes}.")
            
        # Guardar caché de metadatos procesados de este mes
        try:
            with open(metadataF, "w", encoding="utf-8") as mf:
                json.dump(metadata, mf, indent=4, ensure_ascii=False)
        except Exception as e:
            print(f"Error guardando metadatos para {mes}: {e}")
            
        t_month_end = time.time()
        print(f"[Timer] Sincronización del mes {mes} terminada. Tiempo total: {t_month_end - t_month_start:.2f} segundos.")
            
    t_global_end = time.time()
    print(f"[Timer Global] Sincronización de Drive completada. Tiempo total acumulado: {t_global_end - t_global_start:.2f} segundos.")
    return errors, sync_log

if __name__ == "__main__":
    res = obtener_hojas()