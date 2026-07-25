from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from typing import List, Optional
from datetime import datetime
import io
import csv
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment

from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.pagesizes import letter, landscape

from app.database import get_db, get_month_dates
from app.dependencies import DISPONIBLE_STATUSES

def format_agil_month_ranges(records: List[tuple], highlight_html: bool = False) -> str:
    """
    records: list of tuples (day_int, subnovedad_str, [desc]) sorted by day_int
    Returns string with line breaks for each novelty range:
    - HTML/PDF (<br/>): '<font color="#DC2626"><b>10-15</b></font> (VACACIONES)<br/><font color="#DC2626"><b>22</b></font> (PERMISO)'
    - Excel/Plain (\n): '10-15 (VACACIONES)\n22 (PERMISO)'
    """
    if not records:
        return "-"
    
    ranges = []
    curr_start = records[0][0]
    curr_end = records[0][0]
    curr_nov = records[0][1]
    
    def make_label(start: int, end: int, nov: str) -> str:
        day_str = f"{start:02d}" if start == end else f"{start:02d}-{end:02d}"
        if highlight_html:
            return f'<font color="#DC2626"><b>{day_str}</b></font> ({nov})'
        else:
            return f'{day_str} ({nov})'

    for rec in records[1:]:
        day = rec[0]
        nov = rec[1]
        if day == curr_end + 1 and nov == curr_nov:
            curr_end = day
        else:
            ranges.append(make_label(curr_start, curr_end, curr_nov))
            curr_start = day
            curr_end = day
            curr_nov = nov
            
    ranges.append(make_label(curr_start, curr_end, curr_nov))
        
    sep = "<br/>" if highlight_html else "\n"
    return sep.join(ranges)



router = APIRouter(prefix="/api/exportar", tags=["Exportaciones"])


@router.get("/csv")
def exportar_csv(
    tipo: str = Query(..., description="dia, mes, personal, personal_db, subnovedades, consolidado_mensual"),
    fecha: Optional[str] = Query(None),
    mes: Optional[str] = Query(None),
    cedula: Optional[int] = Query(None),
    subnovedad: Optional[str] = Query(None),
    modo: Optional[str] = Query("letras"),
    db = Depends(get_db)
):
    print(f"\n[CSV REQUEST RECEIVED] tipo={tipo} | fecha={fecha} | mes={mes} | cedula={cedula} | subnovedad={subnovedad} | modo={modo}")
    output = io.StringIO()
    writer = csv.writer(output)
    
    cursor = db.cursor()
    
    if tipo == "dia":
        writer.writerow(["CEDULA", "APELLIDOS Y NOMBRES", "SUBNOVEDAD", "DESCRIPCION", "DESDE", "HASTA", "FECHA REPORTE"])
        query = """
            SELECT p.cedula, p.nombre, sn.nombre, rp.descripcion, rp.fecha_inicio, rp.fecha_final, r.fecha
            FROM REGISTRO_PERSONAL rp
            JOIN PERSONAL p ON rp.id_personal = p.id
            JOIN SUB_NOVEDADES sn ON rp.id_sub_novedad = sn.id
            JOIN REPORTES r ON rp.id_reporte = r.id
        """
        params = []
        where_clauses = []
        if fecha:
            where_clauses.append("r.fecha = %s")
            params.append(fecha)
        elif mes and mes.upper() != "TODOS":
            dates = get_month_dates(mes)
            if dates:
                placeholders = ",".join("%s" for _ in dates)
                where_clauses.append(f"r.fecha IN ({placeholders})")
                params.extend(dates)
            else:
                where_clauses.append("1=0")
                
        if where_clauses:
            query += " WHERE " + " AND ".join(where_clauses)
            
        query += " ORDER BY r.fecha ASC, p.nombre ASC;"
        cursor.execute(query, params)
        for row in cursor.fetchall():
            writer.writerow(list(row))
            
        if fecha:
            filename = f"reporte_detallado_dia_{fecha}.csv"
        elif mes and mes.upper() != "TODOS":
            filename = f"reporte_detallado_mes_{mes}.csv"
        else:
            filename = "reporte_detallado_anual.csv"
        
    elif tipo == "mes" and mes:
        writer.writerow(["FECHA", "TOTAL PERSONAL", "DISPONIBLES", "NOVEDADES", "DISPONIBILIDAD %"])
        dates = get_month_dates(mes)
        for d in dates:
            cursor.execute("SELECT id FROM REPORTES WHERE fecha = %s;", (d,))
            r_row = cursor.fetchone()
            if not r_row:
                continue
            r_id = r_row[0]
            
            cursor.execute("SELECT COUNT(*) FROM REGISTRO_PERSONAL WHERE id_reporte = %s;", (r_id,))
            total = cursor.fetchone()[0]
            
            placeholders = ",".join("%s" for _ in DISPONIBLE_STATUSES)
            cursor.execute(f"""
                SELECT COUNT(*) FROM REGISTRO_PERSONAL 
                WHERE id_reporte = %s AND id_sub_novedad IN (
                    SELECT id FROM SUB_NOVEDADES WHERE nombre IN ({placeholders})
                );
            """, (r_id, *DISPONIBLE_STATUSES))
            disp = cursor.fetchone()[0]
            nov = total - disp
            pct = round((disp / total * 100), 1) if total > 0 else 0.0
            writer.writerow([d, total, disp, nov, pct])
        filename = f"reporte_mensual_{mes}.csv"
        
    elif tipo == "personal" and cedula:
        writer.writerow(["FECHA", "SUBNOVEDAD", "DESCRIPCION", "DESDE", "HASTA"])
        cursor.execute("SELECT id, nombre FROM PERSONAL WHERE cedula = %s;", (cedula,))
        p_row = cursor.fetchone()
        if p_row:
            query = """
                SELECT r.fecha, sn.nombre, rp.descripcion, rp.fecha_inicio, rp.fecha_final
                FROM REGISTRO_PERSONAL rp
                JOIN REPORTES r ON rp.id_reporte = r.id
                JOIN SUB_NOVEDADES sn ON rp.id_sub_novedad = sn.id
                WHERE rp.id_personal = %s
            """
            params = [p_row[0]]
            if mes and mes.upper() != "TODOS":
                dates = get_month_dates(mes)
                if dates:
                    placeholders = ",".join("%s" for _ in dates)
                    query += f" AND r.fecha IN ({placeholders})"
                    params.extend(dates)
                else:
                    query += " AND 1=0"
            if subnovedad:
                query += " AND sn.nombre = %s"
                params.append(subnovedad)
            query += " ORDER BY r.fecha ASC;"
            cursor.execute(query, tuple(params))
            for row in cursor.fetchall():
                writer.writerow(list(row))
        filename = f"historial_cedula_{cedula}.csv"
        
    elif tipo == "personal_db":
        writer.writerow(["CEDULA", "APELLIDOS Y NOMBRES", "ESTADO", "FECHA RETIRO"])
        cursor.execute("""
            SELECT cedula, nombre, CASE WHEN fecha_retiro IS NULL THEN 'ACTIVO' ELSE 'RETIRADO' END as estado, fecha_retiro
            FROM PERSONAL
            ORDER BY nombre ASC;
        """)
        for row in cursor.fetchall():
            writer.writerow(list(row))
        filename = "base_datos_personal.csv"
        
    elif tipo == "subnovedades":
        writer.writerow(["ID", "NOMBRE NOVEDAD"])
        cursor.execute("SELECT id, nombre FROM SUB_NOVEDADES ORDER BY nombre ASC;")
        for row in cursor.fetchall():
            writer.writerow(list(row))
        filename = "catalogo_subnovedades.csv"
        
    elif tipo == "consolidado_mensual":
        is_all_months = not mes or mes.upper() == "TODOS" or mes == ""
        use_letras = (modo == "letras")
        
        if fecha:
            cursor.execute("SELECT id, fecha FROM REPORTES WHERE fecha = %s;", (fecha,))
            reports_db = cursor.fetchall()
        elif is_all_months:
            cursor.execute("SELECT id, fecha FROM REPORTES ORDER BY fecha ASC;")
            reports_db = cursor.fetchall()
        else:
            dates = get_month_dates(mes)
            if not dates:
                raise HTTPException(status_code=400, detail="No hay reportes para el mes especificado.")
            placeholders = ",".join("%s" for _ in dates)
            cursor.execute(f"SELECT id, fecha FROM REPORTES WHERE fecha IN ({placeholders}) ORDER BY fecha ASC;", dates)
            reports_db = cursor.fetchall()
            
        report_ids = [r[0] for r in reports_db]
        report_dates = [r[1] for r in reports_db]
        
        if fecha:
            headers = ["CEDULA", "INTEGRANTE", f"FECHA ({fecha})"]
        elif is_all_months:
            headers = ["CEDULA", "INTEGRANTE"] + [f"{d.split('-')[2]}/{d.split('-')[1]}" for d in report_dates]
        else:
            headers = ["CEDULA", "INTEGRANTE"] + [f"Día {d.split('-')[2]}" for d in report_dates]
        writer.writerow(headers)
        
        if report_ids:
            rep_placeholders = ",".join("%s" for _ in report_ids)
            query = f"""
                SELECT p.cedula, p.nombre, p.fecha_retiro, r.fecha as report_fecha, rp.id_reporte, sn.nombre as subnovedad
                FROM REGISTRO_PERSONAL rp
                JOIN PERSONAL p ON rp.id_personal = p.id
                JOIN REPORTES r ON rp.id_reporte = r.id
                JOIN SUB_NOVEDADES sn ON rp.id_sub_novedad = sn.id
                WHERE rp.id_reporte IN ({rep_placeholders})
            """
            params = list(report_ids)
            if cedula:
                query += " AND p.cedula = %s"
                params.append(cedula)
            if subnovedad:
                query += " AND sn.nombre = %s"
                params.append(subnovedad)
            query += " ORDER BY p.nombre ASC;"
            
            cursor.execute(query, params)
            
            person_map = {}
            for row in cursor.fetchall():
                key = (row[0], row[1], row[2])
                if key not in person_map:
                    person_map[key] = {}
                person_map[key][row[4]] = row[5]
                
            for (cedula_val, nombre_val, f_retiro), reports_dict in sorted(person_map.items(), key=lambda x: x[0][1]):
                row_data = [cedula_val, nombre_val]
                for r_id, r_fecha in reports_db:
                    is_retired = False
                    if f_retiro and r_fecha >= f_retiro:
                        is_retired = True
                        
                    if is_retired:
                        row_data.append("R" if use_letras else "RETIRADO")
                    else:
                        raw_nov = reports_dict.get(r_id, "N/A")
                        if use_letras:
                            if raw_nov in DISPONIBLE_STATUSES:
                                cell_str = "D"
                            elif raw_nov == "N/A":
                                cell_str = "-"
                            else:
                                cell_str = "N"
                        else:
                            cell_str = raw_nov
                        row_data.append(cell_str)
                writer.writerow(row_data)
        filename = f"consolidado_personal_{fecha if fecha else (mes if mes else 'TODOS')}.csv"
        
    else:
        raise HTTPException(status_code=400, detail="Parámetros inválidos para la exportación.")
        
    response = StreamingResponse(io.BytesIO(output.getvalue().encode("utf-8-sig")), media_type="text/csv")
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response

@router.get("/excel")
def exportar_excel(
    tipo: str = Query(...),
    fecha: Optional[str] = Query(None),
    mes: Optional[str] = Query(None),
    cedula: Optional[int] = Query(None),
    subnovedad: Optional[str] = Query(None),
    modo: Optional[str] = Query("letras"),
    db = Depends(get_db)
):
    print(f"\n[EXCEL REQUEST RECEIVED] tipo={tipo} | fecha={fecha} | mes={mes} | cedula={cedula} | subnovedad={subnovedad} | modo={modo}")
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Styles
    title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    normal_font = Font(name="Calibri", size=11)
    
    title_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid") # Dark gray
    header_fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid") # Lighter dark gray
    
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    cursor = db.cursor()
    
    if tipo == "dia":
        ws.title = "Reporte Detallado"
        
        ws.merge_cells("A1:G1")
        if fecha:
            title_text = f"BIMEJ12 — REPORTE DETALLADO DE PERSONAL — DÍA {fecha}"
        elif mes and mes.upper() != "TODOS":
            title_text = f"BIMEJ12 — REPORTE DETALLADO DE PERSONAL — MES DE {mes.upper()}"
        else:
            title_text = "BIMEJ12 — REPORTE DETALLADO DE PERSONAL — ANUAL COMPLETO"
            
        ws["A1"] = title_text
        ws["A1"].font = title_font
        ws["A1"].fill = title_fill
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40
        
        headers = ["CÉDULA", "APELLIDOS Y NOMBRES", "SUBNOVEDAD", "DESCRIPCIÓN", "DESDE", "HASTA", "FECHA REPORTE"]
        ws.append([])
        ws.append(headers)
        ws.row_dimensions[3].height = 25
        
        for col_idx in range(1, 8):
            cell = ws.cell(row=3, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        query = """
            SELECT p.cedula, p.nombre, sn.nombre, rp.descripcion, rp.fecha_inicio, rp.fecha_final, r.fecha
            FROM REGISTRO_PERSONAL rp
            JOIN PERSONAL p ON rp.id_personal = p.id
            JOIN SUB_NOVEDADES sn ON rp.id_sub_novedad = sn.id
            JOIN REPORTES r ON rp.id_reporte = r.id
        """
        params = []
        where_clauses = []
        if fecha:
            where_clauses.append("r.fecha = %s")
            params.append(fecha)
        elif mes and mes.upper() != "TODOS":
            dates = get_month_dates(mes)
            if dates:
                placeholders = ",".join("%s" for _ in dates)
                where_clauses.append(f"r.fecha IN ({placeholders})")
                params.extend(dates)
            else:
                where_clauses.append("1=0")
                
        if where_clauses:
            query += " WHERE " + " AND ".join(where_clauses)
            
        query += " ORDER BY r.fecha ASC, p.nombre ASC;"
        cursor.execute(query, params)
        for row in cursor.fetchall():
            ws.append(list(row))
            
        for r_idx in range(4, ws.max_row + 1):
            ws.row_dimensions[r_idx].height = 20
            for c_idx in range(1, 8):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = normal_font
                cell.border = thin_border
                if c_idx == 1:
                    cell.alignment = Alignment(horizontal="left")
                elif c_idx in (5, 6, 7):
                    cell.alignment = Alignment(horizontal="center")
                    
        filename = f"reporte_detallado_{fecha if fecha else (mes if mes else 'anual')}.xlsx"
        
    elif tipo == "mes" and mes:
        ws.title = f"Resumen {mes}"
        ws.merge_cells("A1:E1")
        ws["A1"] = f"BIMEJ12 — RESUMEN MENSUAL — {mes.upper()}"
        ws["A1"].font = title_font
        ws["A1"].fill = title_fill
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40
        
        headers = ["FECHA", "TOTAL PERSONAL", "DISPONIBLES", "NOVEDADES", "DISPONIBILIDAD %"]
        ws.append([])
        ws.append(headers)
        ws.row_dimensions[3].height = 25
        
        for col_idx in range(1, 6):
            cell = ws.cell(row=3, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        dates = get_month_dates(mes)
        for d in dates:
            cursor.execute("SELECT id FROM REPORTES WHERE fecha = %s;", (d,))
            r_row = cursor.fetchone()
            if not r_row:
                continue
            r_id = r_row[0]
            
            cursor.execute("SELECT COUNT(*) FROM REGISTRO_PERSONAL WHERE id_reporte = %s;", (r_id,))
            total = cursor.fetchone()[0]
            
            placeholders = ",".join("%s" for _ in DISPONIBLE_STATUSES)
            cursor.execute(f"""
                SELECT COUNT(*) FROM REGISTRO_PERSONAL 
                WHERE id_reporte = %s AND id_sub_novedad IN (
                    SELECT id FROM SUB_NOVEDADES WHERE nombre IN ({placeholders})
                );
            """, (r_id, *DISPONIBLE_STATUSES))
            disp = cursor.fetchone()[0]
            nov = total - disp
            pct = round((disp / total * 100), 1) if total > 0 else 0.0
            ws.append([d, total, disp, nov, pct])
            
        for r_idx in range(4, ws.max_row + 1):
            ws.row_dimensions[r_idx].height = 20
            for c_idx in range(1, 6):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = normal_font
                cell.border = thin_border
                if c_idx == 1:
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="right")
                    
        filename = f"reporte_mensual_{mes}.xlsx"
        
    elif tipo == "personal" and cedula:
        cursor.execute("SELECT id, nombre, CASE WHEN fecha_retiro IS NULL THEN 'ACTIVO' ELSE 'RETIRADO' END as estado FROM PERSONAL WHERE cedula = %s;", (cedula,))
        p_row = cursor.fetchone()
        nombre = p_row[1] if p_row else "Desconocido"
        estado = p_row[2] if p_row else "Desconocido"
        
        ws.title = "Historial"
        ws.merge_cells("A1:E1")
        ws["A1"] = f"HISTORIAL INDIVIDUAL: {nombre} ({cedula}) - {estado}"
        ws["A1"].font = title_font
        ws["A1"].fill = title_fill
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40
        
        headers = ["FECHA", "SUBNOVEDAD", "DESCRIPCIÓN", "DESDE", "HASTA"]
        ws.append([])
        ws.append(headers)
        ws.row_dimensions[3].height = 25
        
        for col_idx in range(1, 6):
            cell = ws.cell(row=3, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        if p_row:
            query = """
                SELECT r.fecha, sn.nombre, rp.descripcion, rp.fecha_inicio, rp.fecha_final
                FROM REGISTRO_PERSONAL rp
                JOIN REPORTES r ON rp.id_reporte = r.id
                JOIN SUB_NOVEDADES sn ON rp.id_sub_novedad = sn.id
                WHERE rp.id_personal = %s
            """
            params = [p_row[0]]
            if mes and mes.upper() != "TODOS":
                dates = get_month_dates(mes)
                if dates:
                    placeholders = ",".join("%s" for _ in dates)
                    query += f" AND r.fecha IN ({placeholders})"
                    params.extend(dates)
                else:
                    query += " AND 1=0"
            if subnovedad:
                query += " AND sn.nombre = %s"
                params.append(subnovedad)
            query += " ORDER BY r.fecha ASC;"
            cursor.execute(query, tuple(params))
            for row in cursor.fetchall():
                ws.append(list(row))
                
        for r_idx in range(4, ws.max_row + 1):
            ws.row_dimensions[r_idx].height = 20
            for c_idx in range(1, 6):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = normal_font
                cell.border = thin_border
                if c_idx == 1:
                    cell.alignment = Alignment(horizontal="center")
                elif c_idx in (4, 5):
                    cell.alignment = Alignment(horizontal="center")
                    
        filename = f"historial_personal_{cedula}.xlsx"
        
    elif tipo == "personal_db":
        ws.title = "Base Personal"
        ws.merge_cells("A1:D1")
        ws["A1"] = "BIMEJ12 — BASE DE DATOS DE PERSONAL"
        ws["A1"].font = title_font
        ws["A1"].fill = title_fill
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40
        
        headers = ["CÉDULA", "APELLIDOS Y NOMBRES", "ESTADO", "FECHA RETIRO"]
        ws.append([])
        ws.append(headers)
        ws.row_dimensions[3].height = 25
        
        for col_idx in range(1, 5):
            cell = ws.cell(row=3, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        cursor.execute("""
            SELECT cedula, nombre, CASE WHEN fecha_retiro IS NULL THEN 'ACTIVO' ELSE 'RETIRADO' END as estado, fecha_retiro
            FROM PERSONAL
            ORDER BY nombre ASC;
        """)
        for row in cursor.fetchall():
            ws.append(list(row))
            
        for r_idx in range(4, ws.max_row + 1):
            ws.row_dimensions[r_idx].height = 20
            for c_idx in range(1, 5):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = normal_font
                cell.border = thin_border
                if c_idx == 3:
                    cell.alignment = Alignment(horizontal="center")
                    if cell.value == "ACTIVO":
                        cell.font = Font(name="Calibri", size=11, color="10B981", bold=True)
                    else:
                        cell.font = Font(name="Calibri", size=11, color="EF4444", bold=True)
                elif c_idx == 4:
                    cell.alignment = Alignment(horizontal="center")
        filename = "base_datos_personal.xlsx"
        
    elif tipo == "subnovedades":
        ws.title = "Catálogo Novedades"
        ws.merge_cells("A1:B1")
        ws["A1"] = "BIMEJ12 — CATÁLOGO DE SUBNOVEDADES"
        ws["A1"].font = title_font
        ws["A1"].fill = title_fill
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40
        
        headers = ["ID", "NOMBRE NOVEDAD"]
        ws.append([])
        ws.append(headers)
        ws.row_dimensions[3].height = 25
        
        for col_idx in range(1, 3):
            cell = ws.cell(row=3, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        cursor.execute("SELECT id, nombre FROM SUB_NOVEDADES ORDER BY nombre ASC;")
        for row in cursor.fetchall():
            ws.append(list(row))
            
        for r_idx in range(4, ws.max_row + 1):
            ws.row_dimensions[r_idx].height = 20
            for c_idx in range(1, 3):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = normal_font
                cell.border = thin_border
                if c_idx == 1:
                    cell.alignment = Alignment(horizontal="center")
        filename = "catalogo_subnovedades.xlsx"
        
    elif tipo == "consolidado_mensual":
        is_all_months = not mes or mes.upper() == "TODOS" or mes == ""
        use_letras = (modo == "letras")
        
        ws.title = "Consolidado Completo" if is_all_months else f"Consolidado {mes if mes else ''}"
        
        if fecha:
            cursor.execute("SELECT id, fecha FROM REPORTES WHERE fecha = %s;", (fecha,))
            reports_db = cursor.fetchall()
        elif is_all_months:
            cursor.execute("SELECT fecha FROM REPORTES ORDER BY fecha ASC;")
            dates = [row[0] for row in cursor.fetchall()]
            if not dates:
                raise HTTPException(status_code=400, detail="No hay reportes registrados en el sistema.")
            placeholders = ",".join("%s" for _ in dates)
            cursor.execute(f"SELECT id, fecha FROM REPORTES WHERE fecha IN ({placeholders}) ORDER BY fecha ASC;", dates)
            reports_db = cursor.fetchall()
        else:
            dates = get_month_dates(mes)
            if not dates:
                raise HTTPException(status_code=400, detail="No hay reportes para el mes especificado.")
            placeholders = ",".join("%s" for _ in dates)
            cursor.execute(f"SELECT id, fecha FROM REPORTES WHERE fecha IN ({placeholders}) ORDER BY fecha ASC;", dates)
            reports_db = cursor.fetchall()
            
        report_ids = [r[0] for r in reports_db]
        report_dates = [r[1] for r in reports_db]
        
        num_cols = 2 + len(report_dates)
        col_letter = get_column_letter(num_cols)
        
        ws.merge_cells(f"A1:{col_letter}1")
        if cedula:
            title_text = f"BIMEJ12 — HISTORIAL DE PERSONAL (CC {cedula})"
        else:
            title_text = "BIMEJ12 — CONSOLIDADO DIARIO DE PERSONAL"
            
        if fecha:
            title_text += f" — DÍA {fecha}"
        elif not is_all_months:
            title_text += f" — {mes.upper()}"
        else:
            title_text += " — TODOS LOS MESES"
            
        ws["A1"] = title_text
        ws["A1"].font = title_font
        ws["A1"].fill = title_fill
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40
        
        if fecha:
            headers = ["CÉDULA", "INTEGRANTE", f"FECHA ({fecha})"]
        elif is_all_months:
            headers = ["CÉDULA", "INTEGRANTE"] + [f"{d.split('-')[2]}/{d.split('-')[1]}" for d in report_dates]
        else:
            headers = ["CÉDULA", "INTEGRANTE"] + [f"Día {d.split('-')[2]}" for d in report_dates]
            
        ws.append([])
        ws.append(headers)
        ws.row_dimensions[3].height = 25
        
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        if report_ids:
            rep_placeholders = ",".join("%s" for _ in report_ids)
            query = f"""
                SELECT p.cedula, p.nombre, p.fecha_retiro, r.fecha as report_fecha, rp.id_reporte, sn.nombre as subnovedad
                FROM REGISTRO_PERSONAL rp
                JOIN PERSONAL p ON rp.id_personal = p.id
                JOIN REPORTES r ON rp.id_reporte = r.id
                JOIN SUB_NOVEDADES sn ON rp.id_sub_novedad = sn.id
                WHERE rp.id_reporte IN ({rep_placeholders})
            """
            params = list(report_ids)
            if cedula:
                query += " AND p.cedula = %s"
                params.append(cedula)
            if subnovedad:
                query += " AND sn.nombre = %s"
                params.append(subnovedad)
            query += " ORDER BY p.nombre ASC;"
            
            cursor.execute(query, params)
            
            person_map = {}
            for row in cursor.fetchall():
                key = (row[0], row[1], row[2])
                if key not in person_map:
                    person_map[key] = {}
                person_map[key][row[4]] = row[5]
                
            for (cedula_val, nombre_val, f_retiro), reports_dict in sorted(person_map.items(), key=lambda x: x[0][1]):
                row_data = [cedula_val, nombre_val]
                for r_id, r_fecha in reports_db:
                    is_retired = False
                    if f_retiro and r_fecha >= f_retiro:
                        is_retired = True
                        
                    if is_retired:
                        cell_str = "R" if use_letras else "RETIRADO"
                    else:
                        raw_nov = reports_dict.get(r_id, "N/A")
                        if use_letras:
                            if raw_nov in DISPONIBLE_STATUSES:
                                cell_str = "D"
                            elif raw_nov == "N/A":
                                cell_str = "-"
                            else:
                                cell_str = "N"
                        else:
                            cell_str = raw_nov
                    row_data.append(cell_str)
                ws.append(row_data)
                
        for r_idx in range(4, ws.max_row + 1):
            ws.row_dimensions[r_idx].height = 24 if not use_letras else 20
            for c_idx in range(1, num_cols + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = normal_font
                cell.border = thin_border
                if c_idx >= 3:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    val = str(cell.value)
                    if val in DISPONIBLE_STATUSES or val == "D":
                        cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                        cell.font = Font(name="Calibri", size=9, color="065F46", bold=True)
                    elif val in ("N/A", "-"):
                        cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
                        cell.font = Font(name="Calibri", size=9, color="6B7280")
                    elif val in ("RETIRADO", "R"):
                        cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                        cell.font = Font(name="Calibri", size=9, color="DC2626", bold=True)
                    else:
                        cell.fill = PatternFill(start_color="FFE4E6", end_color="FFE4E6", fill_type="solid")
                        cell.font = Font(name="Calibri", size=8 if not use_letras else 9, color="991B1B", bold=True)
                        
        filename = f"consolidado_personal_{fecha if fecha else (mes if mes else 'TODOS')}.xlsx"


        
    elif tipo == "historial_novedades":
        ws.title = "Historial Novedades"
        ws.merge_cells("A1:G1")
        ws["A1"] = "BIMEJ12 — HISTORIAL COMPLETO DE NOVEDADES"
        ws["A1"].font = title_font
        ws["A1"].fill = title_fill
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40
        
        headers = ["CÉDULA", "APELLIDOS Y NOMBRES", "SUBNOVEDAD", "DESCRIPCIÓN", "DESDE", "HASTA", "FECHA REPORTE"]
        ws.append([])
        ws.append(headers)
        ws.row_dimensions[3].height = 25
        
        for col_idx in range(1, 8):
            cell = ws.cell(row=3, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        cursor.execute("""
            SELECT p.cedula, p.nombre, sn.nombre, rp.descripcion, rp.fecha_inicio, rp.fecha_final, r.fecha
            FROM REGISTRO_PERSONAL rp
            JOIN PERSONAL p ON rp.id_personal = p.id
            JOIN SUB_NOVEDADES sn ON rp.id_sub_novedad = sn.id
            JOIN REPORTES r ON rp.id_reporte = r.id
            ORDER BY r.fecha ASC, p.nombre ASC;
        """)
        for row in cursor.fetchall():
            ws.append(list(row))
            
        for r_idx in range(4, ws.max_row + 1):
            ws.row_dimensions[r_idx].height = 20
            for c_idx in range(1, 8):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = normal_font
                cell.border = thin_border
                if c_idx == 1:
                    cell.alignment = Alignment(horizontal="left")
                elif c_idx in (5, 6, 7):
                    cell.alignment = Alignment(horizontal="center")
        filename = "historial_completo_novedades.xlsx"
        
    elif tipo == "agil":
        is_all_months = not mes or mes.upper() == "TODOS" or mes == ""
        
        ws.title = "Exportación Ágil"
        ws.merge_cells("A1:G1")
        
        if fecha:
            title_text = f"BIMEJ12 — EXPORTACIÓN ÁGIL DE NOVEDADES — FECHA: {fecha}"
        elif not is_all_months:
            title_text = f"BIMEJ12 — EXPORTACIÓN ÁGIL DE NOVEDADES — MES DE {mes.upper()}"
        else:
            title_text = "BIMEJ12 — EXPORTACIÓN ÁGIL ANUAL DE NOVEDADES (TODOS LOS MESES)"
            
        if cedula:
            title_text += f" (CC {cedula})"
            
        ws["A1"] = title_text
        ws["A1"].font = title_font
        ws["A1"].fill = title_fill
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40

        ws.merge_cells("A2:G2")
        ws["A2"] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Consolidado Exclusivo de Novedades (Excluye Disponibilidad)"
        ws["A2"].font = Font(name="Calibri", size=9, italic=True, color="64748B")
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 20

        placeholders_disp = ",".join("%s" for _ in DISPONIBLE_STATUSES)
        query = f"""
            SELECT p.cedula, p.nombre, r.fecha, sn.nombre as subnovedad, rp.descripcion
            FROM REGISTRO_PERSONAL rp
            JOIN PERSONAL p ON rp.id_personal = p.id
            JOIN REPORTES r ON rp.id_reporte = r.id
            JOIN SUB_NOVEDADES sn ON rp.id_sub_novedad = sn.id
            WHERE sn.nombre NOT IN ({placeholders_disp})
        """
        params = list(DISPONIBLE_STATUSES)
        
        if cedula:
            query += " AND p.cedula = %s"
            params.append(cedula)
        if fecha:
            query += " AND r.fecha = %s"
            params.append(fecha)
        elif not is_all_months:
            dates = get_month_dates(mes)
            if dates:
                pl = ",".join("%s" for _ in dates)
                query += f" AND r.fecha IN ({pl})"
                params.extend(dates)
                
        query += " ORDER BY p.nombre ASC, r.fecha ASC;"
        cursor.execute(query, params)
        rows = cursor.fetchall()

        if fecha:
            headers = ["CÉDULA", "INTEGRANTE", "NOVEDAD", "DESCRIPCIÓN", "FECHA"]
            ws.append([])
            ws.append(headers)
            ws.row_dimensions[4].height = 25
            for col_idx in range(1, 6):
                cell = ws.cell(row=4, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            for row in rows:
                ws.append([row[0], row[1], row[3], row[4] or "-", row[2]])
            for r_idx in range(5, ws.max_row + 1):
                ws.row_dimensions[r_idx].height = 20
                for c_idx in range(1, 6):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.font = normal_font
                    cell.border = thin_border
                    if c_idx in (1, 5):
                        cell.alignment = Alignment(horizontal="center")
            filename = f"exportacion_agil_{fecha}.xlsx"

        elif not is_all_months:
            headers = ["CÉDULA", "INTEGRANTE", f"RESUMEN DE NOVEDADES - {mes.upper()}"]
            ws.append([])
            ws.append(headers)
            ws.row_dimensions[4].height = 25
            for col_idx in range(1, 4):
                cell = ws.cell(row=4, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                
            from collections import defaultdict
            person_novs = defaultdict(list)
            for r in rows:
                c_num, p_name, r_date, subnov, desc = r
                day_num = int(r_date.split('-')[2])
                person_novs[(c_num, p_name)].append((day_num, subnov, desc))
                
            for (c_num, p_name), recs in sorted(person_novs.items(), key=lambda x: x[0][1]):
                summary_str = format_agil_month_ranges(recs)
                ws.append([c_num, p_name, summary_str])
                current_r = ws.max_row
                
                obs = [f"• Día {d:02d} ({sn}): {desc.strip()}" for d, sn, desc in recs if desc and str(desc).strip()]
                if obs:
                    ws.cell(row=current_r, column=3).comment = Comment("DESCRIPCIÓN DE NOVEDADES:\n" + "\n".join(obs), "BIMEJ12")
                
            for r_idx in range(5, ws.max_row + 1):
                cell_val = ws.cell(row=r_idx, column=3).value or ""
                num_lines = str(cell_val).count('\n') + 1
                ws.row_dimensions[r_idx].height = max(22, num_lines * 18)
                for c_idx in range(1, 4):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.font = normal_font
                    cell.border = thin_border
                    if c_idx == 1:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif c_idx == 2:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    elif c_idx == 3:
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            filename = f"exportacion_agil_{mes}.xlsx"

        else:
            active_m_codes = set(r[2].split('-')[1] for r in rows)
            if not active_m_codes:
                cursor.execute("SELECT DISTINCT to_char(to_date(fecha, 'YYYY-MM-DD'), 'MM') FROM REPORTES;")
                active_m_codes = set(row[0] for row in cursor.fetchall())
                
            all_month_tuples = [
                ('01', 'ENERO'), ('02', 'FEBRERO'), ('03', 'MARZO'), ('04', 'ABRIL'),
                ('05', 'MAYO'), ('06', 'JUNIO'), ('07', 'JULIO'), ('08', 'AGOSTO'),
                ('09', 'SEPTIEMBRE'), ('10', 'OCTUBRE'), ('11', 'NOVIEMBRE'), ('12', 'DICIEMBRE')
            ]
            month_names_dict = [m for m in all_month_tuples if m[0] in active_m_codes]
            headers = ["CÉDULA", "INTEGRANTE"] + [m_name for _, m_name in month_names_dict]
            ws.append([])
            ws.append(headers)
            ws.row_dimensions[4].height = 25

            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=4, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                
            from collections import defaultdict
            person_months = defaultdict(lambda: defaultdict(list))
            for r in rows:
                c_num, p_name, r_date, subnov, desc = r
                m_code = r_date.split('-')[1]
                day_num = int(r_date.split('-')[2])
                person_months[(c_num, p_name)][m_code].append((day_num, subnov, desc))
                
            for (c_num, p_name), m_dict in sorted(person_months.items(), key=lambda x: x[0][1]):
                row_data = [c_num, p_name]
                month_comments = {}
                for idx, (m_code, _) in enumerate(month_names_dict):
                    recs = m_dict.get(m_code, [])
                    row_data.append(format_agil_month_ranges(recs))
                    
                    obs = [f"• Día {d:02d} ({sn}): {desc.strip()}" for d, sn, desc in recs if desc and str(desc).strip()]
                    if obs:
                        col_number = 3 + idx
                        month_comments[col_number] = "DESCRIPCIÓN DE NOVEDADES:\n" + "\n".join(obs)
                        
                ws.append(row_data)
                current_r = ws.max_row
                for col_num, comment_txt in month_comments.items():
                    ws.cell(row=current_r, column=col_num).comment = Comment(comment_txt, "BIMEJ12")

            for r_idx in range(5, ws.max_row + 1):
                max_lines = 1
                for c_idx in range(3, len(headers) + 1):
                    val = str(ws.cell(row=r_idx, column=c_idx).value or "")
                    max_lines = max(max_lines, val.count('\n') + 1)
                ws.row_dimensions[r_idx].height = max(22, max_lines * 18)
                
                for c_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.font = normal_font
                    cell.border = thin_border
                    if c_idx == 1:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif c_idx == 2:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            filename = f"exportacion_agil_anual_{cedula if cedula else 'todos'}.xlsx"


        
    else:
        raise HTTPException(status_code=400, detail="Parámetros inválidos")
        
    # Auto-adjust column widths
    if tipo == "consolidado_mensual":
        if not (cedula and (not mes or mes.upper() == "TODOS" or mes == "")):
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 28
            for col_idx in range(3, num_cols + 1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = 8
    else:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row == 1:
                    continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        
    out_file = io.BytesIO()
    wb.save(out_file)
    out_file.seek(0)
    
    response = StreamingResponse(out_file, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response

@router.get("/pdf")
def exportar_pdf(
    tipo: str = Query(...),
    fecha: Optional[str] = Query(None),
    mes: Optional[str] = Query(None),
    cedula: Optional[int] = Query(None),
    subnovedad: Optional[str] = Query(None),
    modo: Optional[str] = Query("letras"),
    db = Depends(get_db)
):
    print(f"\n==========================================")
    print(f"[PDF REQUEST RECEIVED] tipo={tipo} | fecha={fecha} | mes={mes} | cedula={cedula} | subnovedad={subnovedad} | modo={modo}")
    print(f"==========================================")

    pdf_buffer = io.BytesIO()
    
    # Select document layout based on export type
    doc_layout = landscape(letter) if tipo in ("dia", "mes", "consolidado_mensual") else letter
    doc = SimpleDocTemplate(
        pdf_buffer,
        pagesize=doc_layout,
        leftMargin=36,
        rightMargin=36,
        topMargin=36,
        bottomMargin=36
    )
    
    styles = getSampleStyleSheet()
    
    # Custom Styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=colors.HexColor('#0F172A'),
        alignment=1, # Center
        spaceAfter=15
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=11,
        textColor=colors.HexColor('#475569'),
        alignment=1, # Center
        spaceAfter=20
    )
    
    th_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.white,
        alignment=1 # Center
    )
    
    td_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        textColor=colors.HexColor('#1E293B'),
        alignment=0 # Left
    )
    
    story = []
    cursor = db.cursor()
    
    if tipo == "dia":
        if fecha:
            story.append(Paragraph(f"BIMEJ12 — REPORTE DETALLADO DE PERSONAL — DÍA {fecha}", title_style))
            story.append(Paragraph(f"Fecha del Reporte: {fecha} | Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style))
        elif mes and mes.upper() != "TODOS":
            story.append(Paragraph(f"BIMEJ12 — REPORTE DETALLADO DE PERSONAL — MES DE {mes.upper()}", title_style))
            story.append(Paragraph(f"Mes: {mes.upper()} | Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style))
        else:
            story.append(Paragraph(f"BIMEJ12 — REPORTE DETALLADO DE PERSONAL — ANUAL COMPLETO", title_style))
            story.append(Paragraph(f"Historial General Completo | Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style))
            
        headers = [
            Paragraph("CÉDULA", th_style),
            Paragraph("APELLIDOS Y NOMBRES", th_style),
            Paragraph("SUBNOVEDAD", th_style),
            Paragraph("DESCRIPCIÓN", th_style),
            Paragraph("DESDE", th_style),
            Paragraph("HASTA", th_style),
            Paragraph("FECHA", th_style)
        ]
        
        data = [headers]
        
        query = """
            SELECT p.cedula, p.nombre, sn.nombre, rp.descripcion, rp.fecha_inicio, rp.fecha_final, r.fecha
            FROM REGISTRO_PERSONAL rp
            JOIN PERSONAL p ON rp.id_personal = p.id
            JOIN SUB_NOVEDADES sn ON rp.id_sub_novedad = sn.id
            JOIN REPORTES r ON rp.id_reporte = r.id
        """
        params = []
        where_clauses = []
        if fecha:
            where_clauses.append("r.fecha = %s")
            params.append(fecha)
        elif mes and mes.upper() != "TODOS":
            dates = get_month_dates(mes)
            if dates:
                placeholders = ",".join("%s" for _ in dates)
                where_clauses.append(f"r.fecha IN ({placeholders})")
                params.extend(dates)
            else:
                where_clauses.append("1=0")
                
        if where_clauses:
            query += " WHERE " + " AND ".join(where_clauses)
            
        query += " ORDER BY r.fecha ASC, p.nombre ASC;"
        cursor.execute(query, params)
        
        for row in cursor.fetchall():
            data.append([
                Paragraph(str(row[0]), td_style),
                Paragraph(row[1], td_style),
                Paragraph(row[2], td_style),
                Paragraph(row[3] or "", td_style),
                Paragraph(row[4] or "-", td_style),
                Paragraph(row[5] or "-", td_style),
                Paragraph(row[6], td_style)
            ])
            
        col_widths = [65, 160, 105, 175, 70, 70, 75]
        
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')])
        ]))
        story.append(t)
        filename = f"reporte_detallado_{fecha if fecha else (mes if mes else 'anual')}.pdf"
        
    elif tipo == "mes" and mes:
        story.append(Paragraph(f"BIMEJ12 — RESUMEN MENSUAL DE DISPONIBILIDAD", title_style))
        story.append(Paragraph(f"Mes: {mes.upper()} | Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style))
        
        headers = [
            Paragraph("FECHA", th_style),
            Paragraph("TOTAL PERSONAL", th_style),
            Paragraph("DISPONIBLES", th_style),
            Paragraph("EN NOVEDADES", th_style),
            Paragraph("DISPONIBILIDAD %", th_style)
        ]
        
        data = [headers]
        
        dates = get_month_dates(mes)
        for d in dates:
            cursor.execute("SELECT id FROM REPORTES WHERE fecha = %s;", (d,))
            r_row = cursor.fetchone()
            if not r_row:
                continue
            r_id = r_row[0]
            
            cursor.execute("SELECT COUNT(*) FROM REGISTRO_PERSONAL WHERE id_reporte = %s;", (r_id,))
            total = cursor.fetchone()[0]
            
            placeholders = ",".join("%s" for _ in DISPONIBLE_STATUSES)
            cursor.execute(f"""
                SELECT COUNT(*) FROM REGISTRO_PERSONAL 
                WHERE id_reporte = %s AND id_sub_novedad IN (
                    SELECT id FROM SUB_NOVEDADES WHERE nombre IN ({placeholders})
                );
            """, (r_id, *DISPONIBLE_STATUSES))
            disp = cursor.fetchone()[0]
            nov = total - disp
            pct = round((disp / total * 100), 1) if total > 0 else 0.0
            
            data.append([
                Paragraph(d, td_style),
                Paragraph(str(total), td_style),
                Paragraph(str(disp), td_style),
                Paragraph(str(nov), td_style),
                Paragraph(f"{pct}%", td_style)
            ])
            
        col_widths = [144, 144, 144, 144, 144] # 720 total
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')])
        ]))
        story.append(t)
        filename = f"reporte_mensual_{mes}.pdf"

    elif tipo == "personal" and cedula:
        cursor.execute("SELECT id, nombre, CASE WHEN fecha_retiro IS NULL THEN 'ACTIVO' ELSE 'RETIRADO' END as estado, fecha_retiro FROM PERSONAL WHERE cedula = %s;", (cedula,))
        p_row = cursor.fetchone()
        if not p_row:
            raise HTTPException(status_code=404, detail="Personal no encontrado")
            
        p_id, nombre, estado, fecha_retiro = p_row[0], p_row[1], p_row[2], p_row[3]
        
        story.append(Paragraph(f"HISTORIAL DE PERSONAL INDIVIDUAL", title_style))
        story.append(Paragraph(f"Integrante: {nombre} | Cédula: {cedula} | Estado: {estado} " + (f"| Fecha Retiro: {fecha_retiro}" if fecha_retiro else ""), subtitle_style))
        
        cursor.execute("SELECT COUNT(*) FROM REGISTRO_PERSONAL WHERE id_personal = %s;", (p_id,))
        total_dias = cursor.fetchone()[0]
        
        placeholders = ",".join("%s" for _ in DISPONIBLE_STATUSES)
        cursor.execute(f"""
            SELECT COUNT(*) FROM REGISTRO_PERSONAL 
            WHERE id_personal = %s AND id_sub_novedad IN (
                SELECT id FROM SUB_NOVEDADES WHERE nombre IN ({placeholders})
            );
        """, (p_id, *DISPONIBLE_STATUSES))
        disp_dias = cursor.fetchone()[0]
        nov_dias = total_dias - disp_dias
        disp_pct = round((disp_dias / total_dias * 100), 1) if total_dias > 0 else 0.0
        
        stats_data = [
            [Paragraph("<b>Total Días Registrados:</b>", td_style), Paragraph(str(total_dias), td_style),
             Paragraph("<b>Días Disponibles:</b>", td_style), Paragraph(f"{disp_dias} ({disp_pct}%)", td_style)],
            [Paragraph("<b>Días en Novedades:</b>", td_style), Paragraph(f"{nov_dias} ({round(100.0 - disp_pct, 1)}%)", td_style),
             Paragraph("<b>Estado Actual:</b>", td_style), Paragraph(estado, td_style)]
        ]
        stats_table = Table(stats_data, colWidths=[135, 135, 135, 135])
        stats_table.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F8FAFC')),
            ('PADDING', (0,0), (-1,-1), 6)
        ]))
        story.append(stats_table)
        story.append(Spacer(1, 15))
        
        headers = [
            Paragraph("FECHA", th_style),
            Paragraph("SUBNOVEDAD", th_style),
            Paragraph("DESCRIPCIÓN", th_style),
            Paragraph("DESDE", th_style),
            Paragraph("HASTA", th_style)
        ]
        
        data = [headers]
        
        query = """
            SELECT r.fecha, sn.nombre, rp.descripcion, rp.fecha_inicio, rp.fecha_final
            FROM REGISTRO_PERSONAL rp
            JOIN REPORTES r ON rp.id_reporte = r.id
            JOIN SUB_NOVEDADES sn ON rp.id_sub_novedad = sn.id
            WHERE rp.id_personal = %s
        """
        params = [p_id]
        if mes and mes.upper() != "TODOS":
            dates = get_month_dates(mes)
            if dates:
                placeholders = ",".join("%s" for _ in dates)
                query += f" AND r.fecha IN ({placeholders})"
                params.extend(dates)
            else:
                query += " AND 1=0"
        if subnovedad:
            query += " AND sn.nombre = %s"
            params.append(subnovedad)
        query += " ORDER BY r.fecha ASC;"
        cursor.execute(query, tuple(params))
        
        for row in cursor.fetchall():
            data.append([
                Paragraph(row[0], td_style),
                Paragraph(row[1], td_style),
                Paragraph(row[2] or "", td_style),
                Paragraph(row[3] or "-", td_style),
                Paragraph(row[4] or "-", td_style)
            ])
            
        col_widths = [75, 110, 195, 80, 80]
        
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')])
        ]))
        story.append(t)
        filename = f"historial_personal_{cedula}.pdf"
        
    elif tipo == "personal_db":
        doc_layout = letter
        doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=doc_layout,
            leftMargin=36,
            rightMargin=36,
            topMargin=36,
            bottomMargin=36
        )
        
        story.append(Paragraph("BIMEJ12 — BASE DE DATOS GENERAL DE PERSONAL", title_style))
        story.append(Paragraph(f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style))
        
        headers = [
            Paragraph("CÉDULA", th_style),
            Paragraph("APELLIDOS Y NOMBRES", th_style),
            Paragraph("ESTADO", th_style),
            Paragraph("FECHA RETIRO", th_style)
        ]
        data = [headers]
        
        cursor.execute("""
            SELECT cedula, nombre, CASE WHEN fecha_retiro IS NULL THEN 'ACTIVO' ELSE 'RETIRADO' END as estado, fecha_retiro
            FROM PERSONAL
            ORDER BY nombre ASC;
        """)
        
        active_style = ParagraphStyle('ActCell', parent=td_style, textColor=colors.HexColor('#10B981'), fontName='Helvetica-Bold')
        ret_style = ParagraphStyle('RetCell', parent=td_style, textColor=colors.HexColor('#EF4444'), fontName='Helvetica-Bold')
        
        for row in cursor.fetchall():
            est_text = row[2]
            est_p = Paragraph(est_text, active_style) if est_text == "ACTIVO" else Paragraph(est_text, ret_style)
            data.append([
                Paragraph(str(row[0]), td_style),
                Paragraph(row[1], td_style),
                est_p,
                Paragraph(row[3] or "-", td_style)
            ])
            
        col_widths = [90, 230, 110, 110]
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')])
        ]))
        story.append(t)
        filename = "base_datos_personal.pdf"

    elif tipo == "subnovedades":
        doc_layout = letter
        doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=doc_layout,
            leftMargin=54,
            rightMargin=54,
            topMargin=54,
            bottomMargin=54
        )
        
        story.append(Paragraph("BIMEJ12 — CATÁLOGO DE SUBNOVEDADES", title_style))
        story.append(Paragraph(f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style))
        
        headers = [
            Paragraph("ID", th_style),
            Paragraph("NOMBRE DE LA SUBNOVEDAD", th_style)
        ]
        data = [headers]
        
        cursor.execute("SELECT id, nombre FROM SUB_NOVEDADES ORDER BY nombre ASC;")
        for row in cursor.fetchall():
            data.append([
                Paragraph(str(row[0]), td_style),
                Paragraph(row[1], td_style)
            ])
            
        col_widths = [100, 404]
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')])
        ]))
        story.append(t)
        filename = "catalogo_subnovedades.pdf"
        
    elif tipo == "consolidado_mensual":
        is_all_months = not mes or mes.upper() == "TODOS" or mes == ""
        use_letras = (modo == "letras")
        
        doc_layout = landscape(letter)
        doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=doc_layout,
            leftMargin=36,
            rightMargin=36,
            topMargin=36,
            bottomMargin=36
        )

        if fecha:
            pdf_title = f"BIMEJ12 — CONSOLIDADO DIARIO DE PERSONAL — DÍA {fecha}"
            story.append(Paragraph(pdf_title, title_style))
            story.append(Paragraph(f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style))
            cursor.execute("SELECT id, fecha FROM REPORTES WHERE fecha = %s;", (fecha,))
            reports_db = cursor.fetchall()
        elif is_all_months:
            pdf_title = "BIMEJ12 — CONSOLIDADO DIARIO DE PERSONAL — TODOS LOS MESES"
            story.append(Paragraph(pdf_title, title_style))
            story.append(Paragraph(f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')} | D = Disponible, N = Novedad, R = Retirado, - = Sin Registro", subtitle_style))
            cursor.execute("SELECT id, fecha FROM REPORTES ORDER BY fecha ASC;")
            reports_db = cursor.fetchall()
        else:
            pdf_title = f"BIMEJ12 — CONSOLIDADO DIARIO DE PERSONAL — {mes.upper()}"
            story.append(Paragraph(pdf_title, title_style))
            mode_desc = "D = Disponible, N = Novedad, R = Retirado" if use_letras else "Detalle Completo de Novedades"
            story.append(Paragraph(f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')} | {mode_desc}", subtitle_style))
            
            dates = get_month_dates(mes)
            if not dates:
                raise HTTPException(status_code=400, detail="No hay reportes para el mes especificado.")
            placeholders = ",".join("%s" for _ in dates)
            cursor.execute(f"SELECT id, fecha FROM REPORTES WHERE fecha IN ({placeholders}) ORDER BY fecha ASC;", dates)
            reports_db = cursor.fetchall()

        report_ids = [r[0] for r in reports_db]
        print(f"[PDF DEBUG] Se encontraron {len(reports_db)} reportes en BD (report_ids: {report_ids[:5]}...)")
        
        # Styles for cells
        p_disp_style = ParagraphStyle('PDispG', parent=td_style, fontSize=5 if not use_letras else 4.5, leading=6, textColor=colors.HexColor('#065F46'), fontName='Helvetica-Bold', alignment=1)
        p_nov_style = ParagraphStyle('PNovG', parent=td_style, fontSize=5 if not use_letras else 4.5, leading=6, textColor=colors.HexColor('#92400E'), fontName='Helvetica-Bold', alignment=1)
        p_na_style = ParagraphStyle('PNAG', parent=td_style, fontSize=5, leading=6, textColor=colors.HexColor('#9CA3AF'), alignment=1)
        p_ret_style = ParagraphStyle('PRetG', parent=td_style, fontSize=4.5, leading=6, textColor=colors.HexColor('#7F1D1D'), fontName='Helvetica-Bold', alignment=1)
        month_title_style = ParagraphStyle('MonthTitle', parent=subtitle_style, fontSize=11, leading=14, textColor=colors.HexColor('#06B6D4'), fontName='Helvetica-Bold', spaceBefore=10, spaceAfter=5)

        if report_ids:
            rep_placeholders = ",".join("%s" for _ in report_ids)
            query = f"""
                SELECT p.cedula, p.nombre, p.fecha_retiro, r.fecha as report_fecha, rp.id_reporte, sn.nombre as subnovedad
                FROM REGISTRO_PERSONAL rp
                JOIN PERSONAL p ON rp.id_personal = p.id
                JOIN REPORTES r ON rp.id_reporte = r.id
                JOIN SUB_NOVEDADES sn ON rp.id_sub_novedad = sn.id
                WHERE rp.id_reporte IN ({rep_placeholders})
            """
            params = list(report_ids)
            if cedula:
                query += " AND p.cedula = %s"
                params.append(cedula)
            if subnovedad:
                query += " AND sn.nombre = %s"
                params.append(subnovedad)
            query += " ORDER BY p.nombre ASC;"
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            print(f"[PDF DEBUG] Filas recuperadas para REGISTRO_PERSONAL: {len(rows)}")
            
            person_map = {}
            for row in rows:
                key = (row[0], row[1], row[2])
                if key not in person_map:
                    person_map[key] = {}
                person_map[key][row[4]] = row[5]
            print(f"[PDF DEBUG] Person map construido con {len(person_map)} personas.")
                
            from collections import defaultdict
            months_grouped = defaultdict(list)
            for r in reports_db:
                m_num = r[1].split('-')[1]
                months_grouped[m_num].append(r)
                
            month_names_dict = {
                '01': 'ENERO', '02': 'FEBRERO', '03': 'MARZO', '04': 'ABRIL',
                ('05'): 'MAYO', '06': 'JUNIO', '07': 'JULIO', '08': 'AGOSTO',
                '09': 'SEPTIEMBRE', '10': 'OCTUBRE', '11': 'NOVIEMBRE', '12': 'DICIEMBRE'
            }

            for m_num, m_reports in sorted(months_grouped.items()):
                m_name = month_names_dict.get(m_num, f"MES {m_num}")
                if is_all_months and not fecha:
                    story.append(Paragraph(f"MES DE {m_name}", month_title_style))
                
                m_dates = [r[1] for r in m_reports]
                headers = [
                    Paragraph("CÉDULA", th_style),
                    Paragraph("INTEGRANTE", th_style)
                ] + [Paragraph(d.split('-')[2], th_style) for d in m_dates]
                
                data = [headers]
                for (c_num, name, f_retiro), reports_dict in sorted(person_map.items(), key=lambda x: x[0][1]):
                    row_data = [
                        Paragraph(str(c_num), td_style),
                        Paragraph(name, td_style)
                    ]
                    for r_id, r_fecha in m_reports:
                        is_retired = False
                        if f_retiro and r_fecha >= f_retiro:
                            is_retired = True
                            
                        if is_retired:
                            row_data.append(Paragraph("R" if use_letras else "RETIRADO", p_ret_style))
                        else:
                            raw_nov = reports_dict.get(r_id, "N/A")
                            if use_letras:
                                if raw_nov in DISPONIBLE_STATUSES:
                                    row_data.append(Paragraph("D", p_disp_style))
                                elif raw_nov == "N/A":
                                    row_data.append(Paragraph("-", p_na_style))
                                else:
                                    row_data.append(Paragraph("N", p_nov_style))
                            else:
                                if raw_nov in DISPONIBLE_STATUSES:
                                    row_data.append(Paragraph(raw_nov, p_disp_style))
                                elif raw_nov == "N/A":
                                    row_data.append(Paragraph("-", p_na_style))
                                else:
                                    row_data.append(Paragraph(raw_nov, p_nov_style))
                    data.append(row_data)
                    
                num_days_col = len(m_dates)
                cedula_width = 45
                min_name_width = 110
                avail_days_width = 720 - cedula_width - min_name_width
                day_col_width = min(max(int(avail_days_width / max(num_days_col, 1)), 12), 120)
                total_days_width = day_col_width * num_days_col
                name_width = 720 - cedula_width - total_days_width
                col_widths = [cedula_width, name_width] + [day_col_width for _ in m_dates]
                
                t = Table(data, colWidths=col_widths, repeatRows=1)
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
                    ('ALIGN', (0,0), (1,-1), 'LEFT'),
                    ('ALIGN', (2,0), (-1,-1), 'CENTER'),
                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                    ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
                    ('TOPPADDING', (0,0), (-1,-1), 1.5 if not use_letras else 2),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 1.5 if not use_letras else 2),
                    ('LEFTPADDING', (0,0), (-1,-1), 1),
                    ('RIGHTPADDING', (0,0), (-1,-1), 1),
                    ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')])
                ]))
                story.append(t)
                story.append(Spacer(1, 10))
                
        filename = f"consolidado_personal_{fecha if fecha else (mes if mes else 'TODOS')}.pdf"



        
    elif tipo == "historial_novedades":
        story.append(Paragraph("BIMEJ12 — HISTORIAL COMPLETO DE NOVEDADES", title_style))
        story.append(Paragraph(f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style))
        
        headers = [
            Paragraph("CÉDULA", th_style),
            Paragraph("APELLIDOS Y NOMBRES", th_style),
            Paragraph("SUBNOVEDAD", th_style),
            Paragraph("DESCRIPCIÓN", th_style),
            Paragraph("DESDE", th_style),
            Paragraph("HASTA", th_style),
            Paragraph("REPORTE", th_style)
        ]
        
        data = [headers]
        
        cursor.execute("""
            SELECT p.cedula, p.nombre, sn.nombre, rp.descripcion, rp.fecha_inicio, rp.fecha_final, r.fecha
            FROM REGISTRO_PERSONAL rp
            JOIN PERSONAL p ON rp.id_personal = p.id
            JOIN SUB_NOVEDADES sn ON rp.id_sub_novedad = sn.id
            JOIN REPORTES r ON rp.id_reporte = r.id
            ORDER BY r.fecha ASC, p.nombre ASC;
        """)
        
        for row in cursor.fetchall():
            data.append([
                Paragraph(str(row[0]), td_style),
                Paragraph(row[1], td_style),
                Paragraph(row[2], td_style),
                Paragraph(row[3] or "", td_style),
                Paragraph(row[4] or "-", td_style),
                Paragraph(row[5] or "-", td_style),
                Paragraph(row[6], td_style)
            ])
            
        col_widths = [70, 150, 100, 160, 80, 80, 80]
        
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')])
        ]))
        filename = "historial_completo_novedades.pdf"
        
    elif tipo == "agil":
        is_all_months = not mes or mes.upper() == "TODOS" or mes == ""
        
        doc_layout = landscape(letter)
        doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=doc_layout,
            leftMargin=36,
            rightMargin=36,
            topMargin=36,
            bottomMargin=36
        )
        
        pdf_title = "BIMEJ12 — EXPORTACIÓN ÁGIL DE NOVEDADES"
        if fecha:
            pdf_title += f" — FECHA: {fecha}"
        elif not is_all_months:
            pdf_title += f" — MES DE {mes.upper()}"
        else:
            pdf_title += " — TODOS LOS MESES (ANUAL)"
            
        if cedula:
            pdf_title += f" (CC {cedula})"
            
        story.append(Paragraph(pdf_title, title_style))
        story.append(Paragraph(f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Consolidado Exclusivo de Novedades (Excluye Disponibilidad)", subtitle_style))
        story.append(Spacer(1, 10))

        placeholders_disp = ",".join("%s" for _ in DISPONIBLE_STATUSES)
        query = f"""
            SELECT p.cedula, p.nombre, r.fecha, sn.nombre as subnovedad, rp.descripcion
            FROM REGISTRO_PERSONAL rp
            JOIN PERSONAL p ON rp.id_personal = p.id
            JOIN REPORTES r ON rp.id_reporte = r.id
            JOIN SUB_NOVEDADES sn ON rp.id_sub_novedad = sn.id
            WHERE sn.nombre NOT IN ({placeholders_disp})
        """
        params = list(DISPONIBLE_STATUSES)
        
        if cedula:
            query += " AND p.cedula = %s"
            params.append(cedula)
        if fecha:
            query += " AND r.fecha = %s"
            params.append(fecha)
        elif not is_all_months:
            dates = get_month_dates(mes)
            if dates:
                pl = ",".join("%s" for _ in dates)
                query += f" AND r.fecha IN ({pl})"
                params.extend(dates)
                
        query += " ORDER BY p.nombre ASC, r.fecha ASC;"
        cursor.execute(query, params)
        rows = cursor.fetchall()

        agil_td_style = ParagraphStyle('AgilTd', parent=td_style, fontSize=7, leading=9)
        agil_th_style = ParagraphStyle('AgilTh', parent=th_style, fontSize=8, leading=10)

        if fecha:
            headers = [
                Paragraph("CÉDULA", agil_th_style),
                Paragraph("INTEGRANTE", agil_th_style),
                Paragraph("NOVEDAD", agil_th_style),
                Paragraph("DESCRIPCIÓN", agil_th_style),
                Paragraph("FECHA", agil_th_style)
            ]
            data = [headers]
            for row in rows:
                data.append([
                    Paragraph(str(row[0]), agil_td_style),
                    Paragraph(row[1], agil_td_style),
                    Paragraph(row[3], agil_td_style),
                    Paragraph(row[4] or "-", agil_td_style),
                    Paragraph(f'<font color="#DC2626"><b>{row[2]}</b></font>', agil_td_style)
                ])
            col_widths = [65, 160, 110, 305, 80]
            filename = f"exportacion_agil_{fecha}.pdf"

        elif not is_all_months:
            headers = [
                Paragraph("CÉDULA", agil_th_style),
                Paragraph("INTEGRANTE", agil_th_style),
                Paragraph(f"RESUMEN DE NOVEDADES ({mes.upper()})", agil_th_style)
            ]
            data = [headers]
            from collections import defaultdict
            person_novs = defaultdict(list)
            for r in rows:
                c_num, p_name, r_date, subnov, desc = r
                day_num = int(r_date.split('-')[2])
                person_novs[(c_num, p_name)].append((day_num, subnov))
                
            for (c_num, p_name), recs in sorted(person_novs.items(), key=lambda x: x[0][1]):
                summary_str = format_agil_month_ranges(recs, highlight_html=True)
                data.append([
                    Paragraph(str(c_num), agil_td_style),
                    Paragraph(p_name, agil_td_style),
                    Paragraph(summary_str, agil_td_style)
                ])
            col_widths = [65, 175, 480]
            filename = f"exportacion_agil_{mes}.pdf"

        else:
            active_m_codes = set(r[2].split('-')[1] for r in rows)
            if not active_m_codes:
                cursor.execute("SELECT DISTINCT to_char(to_date(fecha, 'YYYY-MM-DD'), 'MM') FROM REPORTES;")
                active_m_codes = set(row[0] for row in cursor.fetchall())

            all_month_tuples = [
                ('01', 'ENE'), ('02', 'FEB'), ('03', 'MAR'), ('04', 'ABR'),
                ('05', 'MAY'), ('06', 'JUN'), ('07', 'JUL'), ('08', 'AGO'),
                ('09', 'SEP'), ('10', 'OCT'), ('11', 'NOV'), ('12', 'DIC')
            ]
            month_names_dict = [m for m in all_month_tuples if m[0] in active_m_codes]
            
            headers = [Paragraph("CÉDULA", agil_th_style), Paragraph("INTEGRANTE", agil_th_style)] + [
                Paragraph(m_name, agil_th_style) for _, m_name in month_names_dict
            ]
            data = [headers]
            from collections import defaultdict
            person_months = defaultdict(lambda: defaultdict(list))
            for r in rows:
                c_num, p_name, r_date, subnov, desc = r
                m_code = r_date.split('-')[1]
                day_num = int(r_date.split('-')[2])
                person_months[(c_num, p_name)][m_code].append((day_num, subnov))
                
            for (c_num, p_name), m_dict in sorted(person_months.items(), key=lambda x: x[0][1]):
                row_data = [Paragraph(str(c_num), agil_td_style), Paragraph(p_name, agil_td_style)]
                for m_code, _ in month_names_dict:
                    recs = m_dict.get(m_code, [])
                    summary_str = format_agil_month_ranges(recs, highlight_html=True)
                    row_data.append(Paragraph(summary_str, agil_td_style))
                data.append(row_data)

            month_col_w = max(int(540 / len(month_names_dict)), 46) if month_names_dict else 46
            col_widths = [55, 125] + [month_col_w for _ in month_names_dict]
            filename = f"exportacion_agil_anual_{cedula if cedula else 'todos'}.pdf"

        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
            ('TOPPADDING', (0,0), (-1,-1), 3),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')])
        ]))
        story.append(t)

        
    else:
        raise HTTPException(status_code=400, detail="Parámetros inválidos")
        
    try:
        doc.build(story)
    except Exception as e:
        print(f"Error generando PDF para tipo={tipo}: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error interno al generar PDF: {str(e)}")
        
    pdf_buffer.seek(0)
    
    response = StreamingResponse(pdf_buffer, media_type="application/pdf")
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response
