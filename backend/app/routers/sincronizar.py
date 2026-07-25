import io
import json
import openpyxl
from datetime import datetime
from typing import Optional, List
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from fastapi.responses import StreamingResponse, HTMLResponse
from app.dependencies import DISPONIBLE_STATUSES, get_current_user
from app.database import get_db, get_month_dates
from config.auth import generar_oauth_url, intercambiar_codigo_oauth

router = APIRouter(prefix="/api", tags=["Sincronizar"])

@router.get("/sincronizar/oauth/url")
def obtener_url_oauth(redirect_uri: str = Query(...)):
    """
    Genera la URL oficial de inicio de sesión con Google OAuth.
    """
    try:
        url = generar_oauth_url(redirect_uri)
        return {"auth_url": url}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/sincronizar/oauth/callback")
def callback_oauth(code: str = Query(...), redirect_uri: str = Query(...)):
    """
    Recibe el código de autorización de Google, obtiene y guarda el token.
    """
    try:
        intercambiar_codigo_oauth(code, redirect_uri)
        html_content = """
        <html>
            <head><title>Google Drive Conectado</title></head>
            <body style="font-family: sans-serif; text-align: center; padding-top: 50px; background: #0F172A; color: white;">
                <h1 style="color: #22C55E;">¡Google Drive Conectado con Éxito! 🎉</h1>
                <p>Tu cuenta de Google fue autorizada correctamente.</p>
                <p>Ya puedes volver a la aplicación BIMEH en tu celular o computador.</p>
            </body>
        </html>
        """
        return HTMLResponse(content=html_content)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error en callback OAuth: {str(e)}")

@router.get("/sincronizar/plantilla/{format}")

def descargar_plantilla(format: str):
    """
    Genera y descarga un archivo plantilla para la carga de datos.
    """
    if format == "excel":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Plantilla Reporte"
        
        # Headers
        headers = ["CEDULA", "APELLIDOS Y NOMBRES", "SUBNOVEDAD", "DESCRIPCION", "DESDE", "HASTA"]
        ws.append(headers)
        
        # Sample row
        sample_row = [6804683, "RAMIREZ BOGOYA OMAR", "CDO UNIDAD", "ADMINISTRADOR", "2026-05-07", ""]
        ws.append(sample_row)
        
        # Design formatting
        for col_idx in range(1, 7):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 22
            
        out_file = io.BytesIO()
        wb.save(out_file)
        out_file.seek(0)
        
        response = StreamingResponse(out_file, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response.headers["Content-Disposition"] = "attachment; filename=plantilla_reporte_diario.xlsx"
        return response
        
    elif format == "json":
        sample_json = {
            "2026-05-07": {
                "1": {
                    "CEDULA": 6804683,
                    "APELLIDOS Y NOMBRES": "RAMIREZ BOGOYA OMAR",
                    "SUBNOVEDAD": "CDO UNIDAD",
                    "DESCRIPCION": "ADMINISTRADOR",
                    "DESDE": "2026-05-07",
                    "HASTA": ""
                }
            }
        }
        json_str = json.dumps(sample_json, indent=4, ensure_ascii=False)
        out_file = io.BytesIO(json_str.encode("utf-8-sig"))
        response = StreamingResponse(out_file, media_type="application/json")
        response.headers["Content-Disposition"] = "attachment; filename=plantilla_reporte.json"
        return response
        
    else:
        raise HTTPException(status_code=400, detail="Formato de plantilla no soportado. Use 'excel' o 'json'.")

@router.post("/sincronizar/cargar")
async def cargar_reporte(
    tipo: str = Form(..., description="dia o mes"),
    fecha: Optional[str] = Form(None, description="Fecha YYYY-MM-DD para el tipo dia"),
    mes: Optional[str] = Form(None, description="Nombre del mes para el tipo mes"),
    overwrite: bool = Form(False, description="Booleano para sobreescribir reportes existentes"),
    file: UploadFile = File(...),
    db = Depends(get_db)
):
    cursor = db.cursor()
    filename = file.filename
    contents = await file.read()
    
    # Identify extension
    is_excel = filename.endswith(".xlsx") or filename.endswith(".xls")
    is_json = filename.endswith(".json")
    
    if not is_excel and not is_json:
        raise HTTPException(status_code=400, detail="Formato de archivo inválido. Debe subir un archivo .xlsx o .json.")
        
    # Validation helpers
    required_fields = ["CEDULA", "APELLIDOS Y NOMBRES", "SUBNOVEDAD"]
    
    records_to_sync = {} # Schema: { "YYYY-MM-DD": [ {record}, {record} ] }
    
    if is_excel:
        if tipo == "mes":
            raise HTTPException(status_code=400, detail="El formato Excel solo se puede cargar para reportes de un Día específico.")
            
        if not fecha:
            raise HTTPException(status_code=400, detail="Falta la fecha del reporte para la carga diaria de Excel.")
            
        try:
            wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
            ws = wb.active
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error leyendo el archivo Excel: {str(e)}")
            
        # Find header row
        header_row_idx = None
        headers_map = {}
        
        for r_idx in range(1, min(ws.max_row + 1, 15)):
            row_vals = [str(ws.cell(row=r_idx, column=c_idx).value).strip().upper() for c_idx in range(1, ws.max_column + 1) if ws.cell(row=r_idx, column=c_idx).value is not None]
            if all(f in row_vals for f in required_fields):
                header_row_idx = r_idx
                for c_idx in range(1, ws.max_column + 1):
                    val = str(ws.cell(row=r_idx, column=c_idx).value).strip().upper()
                    headers_map[val] = c_idx
                break
                
        if not header_row_idx:
            raise HTTPException(status_code=400, detail="Estructura incorrecta: Faltan columnas obligatorias (CEDULA, APELLIDOS Y NOMBRES, o SUBNOVEDAD).")
            
        day_records = []
        for r_idx in range(header_row_idx + 1, ws.max_row + 1):
            cedula_val = ws.cell(row=r_idx, column=headers_map["CEDULA"]).value
            nombre_val = ws.cell(row=r_idx, column=headers_map["APELLIDOS Y NOMBRES"]).value
            subnovedad_val = ws.cell(row=r_idx, column=headers_map["SUBNOVEDAD"]).value
            
            if cedula_val is None:
                continue
                
            try:
                # Handle floats and strings safely
                cedula_int = int(float(str(cedula_val).strip()))
            except ValueError:
                continue
                
            if cedula_int <= 0:
                continue
                
            # Optional fields
            descripcion = ""
            if "DESCRIPCION" in headers_map:
                v = ws.cell(row=r_idx, column=headers_map["DESCRIPCION"]).value
                descripcion = str(v).strip() if v is not None else ""
                
            desde = ""
            if "DESDE" in headers_map:
                v = ws.cell(row=r_idx, column=headers_map["DESDE"]).value
                desde = str(v).strip() if v is not None else ""
                
            hasta = ""
            if "HASTA" in headers_map:
                v = ws.cell(row=r_idx, column=headers_map["HASTA"]).value
                hasta = str(v).strip() if v is not None else ""
                
            day_records.append({
                "CEDULA": cedula_int,
                "APELLIDOS Y NOMBRES": str(nombre_val).strip().upper(),
                "SUBNOVEDAD": str(subnovedad_val).strip().upper() if subnovedad_val else "SIN NOVEDAD",
                "DESCRIPCION": descripcion,
                "DESDE": desde,
                "HASTA": hasta
            })
            
        records_to_sync[fecha] = day_records
        
    elif is_json:
        try:
            data = json.loads(contents.decode("utf-8-sig"))
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error leyendo el archivo JSON: {str(e)}")
            
        # Distinguish between daily flat object and nested monthly object
        # If flat object (e.g. { "1": { "CEDULA": 123... } })
        # We check keys. If they are numbers/counters, it's a daily flat JSON
        sample_key = list(data.keys())[0] if data else ""
        
        is_daily_flat = not (len(sample_key) == 10 and sample_key.count("-") == 2)
        
        if is_daily_flat:
            if tipo == "mes":
                raise HTTPException(status_code=400, detail="Estructura de JSON inválida para la carga mensual. Se esperaba un objeto organizado por fechas.")
            if not fecha:
                raise HTTPException(status_code=400, detail="Falta la fecha del reporte para la carga diaria de JSON.")
                
            day_records = []
            for k, record in data.items():
                if not all(field in record for field in required_fields):
                    raise HTTPException(status_code=400, detail="Estructura incorrecta en registros de JSON: Faltan campos obligatorios (CEDULA, APELLIDOS Y NOMBRES, o SUBNOVEDAD).")
                day_records.append({
                    "CEDULA": int(record["CEDULA"]),
                    "APELLIDOS Y NOMBRES": str(record["APELLIDOS Y NOMBRES"]).strip().upper(),
                    "SUBNOVEDAD": str(record["SUBNOVEDAD"]).strip().upper(),
                    "DESCRIPCION": str(record.get("DESCRIPCION", "")),
                    "DESDE": str(record.get("DESDE", "")),
                    "HASTA": str(record.get("HASTA", ""))
                })
            records_to_sync[fecha] = day_records
            
        else: # Monthly nested object (e.g. { "2026-05-07": { "1": { ... } } })
            if tipo == "dia":
                # Extract only selected fecha if it exists in JSON
                if fecha not in data:
                    raise HTTPException(status_code=400, detail=f"La fecha seleccionada '{fecha}' no se encuentra en el archivo JSON subido.")
                day_data = data[fecha]
                day_records = []
                for k, record in day_data.items():
                    if not all(field in record for field in required_fields):
                        raise HTTPException(status_code=400, detail="Estructura incorrecta: Faltan campos obligatorios (CEDULA, APELLIDOS Y NOMBRES, o SUBNOVEDAD).")
                    day_records.append({
                        "CEDULA": int(record["CEDULA"]),
                        "APELLIDOS Y NOMBRES": str(record["APELLIDOS Y NOMBRES"]).strip().upper(),
                        "SUBNOVEDAD": str(record["SUBNOVEDAD"]).strip().upper(),
                        "DESCRIPCION": str(record.get("DESCRIPCION", "")),
                        "DESDE": str(record.get("DESDE", "")),
                        "HASTA": str(record.get("HASTA", ""))
                    })
                records_to_sync[fecha] = day_records
                
            else: # Full monthly load
                if not mes:
                    raise HTTPException(status_code=400, detail="Falta especificar el mes para la carga mensual.")
                    
                # Fetch expected dates for that month in database
                expected_month_dates = get_month_dates(mes)
                
                for d_key, day_data in data.items():
                    if d_key not in expected_month_dates:
                        # Skip dates that don't belong to the selected month
                        continue
                        
                    day_records = []
                    for k, record in day_data.items():
                        if not all(field in record for field in required_fields):
                            raise HTTPException(status_code=400, detail=f"Estructura incorrecta en fecha '{d_key}': Faltan campos obligatorios.")
                        day_records.append({
                            "CEDULA": int(record["CEDULA"]),
                            "APELLIDOS Y NOMBRES": str(record["APELLIDOS Y NOMBRES"]).strip().upper(),
                            "SUBNOVEDAD": str(record["SUBNOVEDAD"]).strip().upper(),
                            "DESCRIPCION": str(record.get("DESCRIPCION", "")),
                            "DESDE": str(record.get("DESDE", "")),
                            "HASTA": str(record.get("HASTA", ""))
                        })
                    records_to_sync[d_key] = day_records

    # Validation: Verify duplication conflicts
    dates_with_conflicts = []
    for d_key in records_to_sync.keys():
        cursor.execute("SELECT id FROM REPORTES WHERE fecha = ?;", (d_key,))
        row = cursor.fetchone()
        if row:
            dates_with_conflicts.append(d_key)
            
    if dates_with_conflicts and not overwrite:
        return {
            "status": "conflict",
            "message": "Ya existen reportes guardados en el sistema para algunas fechas.",
            "conflicts": dates_with_conflicts
        }
        
    # Write to Database
    try:
        for d_key, rows in records_to_sync.items():
            # Delete old registrations if overwriting
            cursor.execute("SELECT id FROM REPORTES WHERE fecha = ?;", (d_key,))
            rep_row = cursor.fetchone()
            if rep_row:
                report_id = rep_row[0]
                cursor.execute("DELETE FROM REGISTRO_PERSONAL WHERE id_reporte = ?;", (report_id,))
            else:
                # Insert report entry
                cursor.execute("INSERT INTO REPORTES (fecha, archivo) VALUES (?, ?);", (d_key, filename))
                cursor.execute("SELECT id FROM REPORTES WHERE fecha = ?;", (d_key,))
                report_id = cursor.fetchone()[0]
                
            # Sincronizar registros
            for r in rows:
                cedula = r["CEDULA"]
                nombre = r["APELLIDOS Y NOMBRES"]
                subnovedad = r["SUBNOVEDAD"]
                descripcion = r["DESCRIPCION"]
                desde_val = r["DESDE"]
                hasta_val = r["HASTA"]
                
                # Normalize subnovedad name
                if not subnovedad:
                    subnovedad = "SIN NOVEDAD"
                else:
                    subnovedad = subnovedad.strip().upper()
                
                # Get or Create Personal
                cursor.execute("SELECT id FROM PERSONAL WHERE cedula = ?;", (cedula,))
                p_row = cursor.fetchone()
                if p_row:
                    personal_id = p_row[0]
                else:
                    cursor.execute("INSERT INTO PERSONAL (cedula, nombre) VALUES (?, ?);", (cedula, nombre))
                    cursor.execute("SELECT id FROM PERSONAL WHERE cedula = ?;", (cedula,))
                    personal_id = cursor.fetchone()[0]
                    
                # Get or Create Subnovedad
                cursor.execute("SELECT id FROM SUB_NOVEDADES WHERE nombre = ?;", (subnovedad,))
                sn_row = cursor.fetchone()
                if sn_row:
                    subnovedad_id = sn_row[0]
                else:
                    cursor.execute("INSERT INTO SUB_NOVEDADES (nombre) VALUES (?);", (subnovedad,))
                    cursor.execute("SELECT id FROM SUB_NOVEDADES WHERE nombre = ?;", (subnovedad,))
                    subnovedad_id = cursor.fetchone()[0]
                    
                # Clean dates for start and end
                fecha_inicio = None
                if desde_val:
                    try:
                        fecha_inicio = desde_val.split()[0]
                    except Exception:
                        pass
                
                fecha_final = None
                if hasta_val:
                    try:
                        fecha_final = hasta_val.split()[0]
                    except Exception:
                        pass
                        
                # Insert Registration
                cursor.execute("""
                    INSERT INTO REGISTRO_PERSONAL 
                    (id_reporte, id_personal, id_sub_novedad, descripcion, fecha_inicio, fecha_final)
                    VALUES (?, ?, ?, ?, ?, ?);
                """, (report_id, personal_id, subnovedad_id, descripcion, fecha_inicio, fecha_final))
                
        db.commit()
    except Exception as ex:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error en base de datos al sincronizar: {str(ex)}")
        
    return {
        "status": "success",
        "message": f"Sincronización completada. Se cargaron reportes para {len(records_to_sync)} fechas operativas."
    }

from pydantic import BaseModel

class DriveSyncRequest(BaseModel):
    tipo: str  # "dia", "dias", "mes" o "todo"
    fecha: Optional[str] = None
    fechas: Optional[list] = None  # Lista de fechas ISO para modo multi-día
    mes: Optional[str] = None
    overwrite: bool = False

def sync_local_jsons_to_db(db, target_month=None, target_date=None, target_dates=None, force_overwrite=False):
    cursor = db.cursor()
    cursor.execute("SELECT fecha FROM REPORTES;")
    existing_dates = {row[0] for row in cursor.fetchall()}
    
    # Pre-cargar tablas maestras en cache local de memoria para evitar consultas N+1 a Neon
    print("Pre-cargando tablas maestras de Neon en memoria para optimización...")
    cursor.execute("SELECT cedula, id FROM PERSONAL;")
    personal_cache = {int(row[0]): int(row[1]) for row in cursor.fetchall()}
    
    cursor.execute("SELECT nombre, id FROM SUB_NOVEDADES;")
    subnovedades_cache = {str(row[0]).strip().upper(): int(row[1]) for row in cursor.fetchall()}
    print(f"Caché cargado: {len(personal_cache)} empleados y {len(subnovedades_cache)} tipos de novedad.")
    
    import os
    from pathlib import Path
    
    months_dir = Path("listadoMeses")
    months_dir.mkdir(parents=True, exist_ok=True)

        
    json_files = [f for f in os.listdir(months_dir) if f.endswith(".json")]
    
    # Si filtramos por un mes en específico, solo procesamos ese archivo JSON
    if target_month:
        json_files = [f"{target_month}.json"]
    elif target_dates:
        # Modo multi-día: determinar los meses involucrados
        reverse_map_db = {
            1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
            5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
            9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
        }
        meses_set = set()
        for td in target_dates:
            try:
                m_num = int(td.split("-")[1])
                mes_name = reverse_map_db.get(m_num)
                if mes_name:
                    meses_set.add(f"{mes_name}.json")
            except Exception:
                pass
        json_files = list(meses_set)
    
    for filename in json_files:
        filepath = months_dir / filename
        if not filepath.exists():
            continue
            
        with open(filepath, "r", encoding="utf-8") as f:
            try:
                data = json.load(f)
            except Exception:
                continue
                
        for date_str, records in data.items():
            # Si filtramos por fechas específicas (multi-día), ignorar las demás
            if target_dates and date_str not in target_dates:
                continue
            # Si filtramos por fecha específica (día único), ignorar las demás
            if target_date and date_str != target_date:
                continue
                
            # Si force_overwrite es True (o si filtramos por esta fecha específica y queremos sobrescribir),
            # eliminamos el reporte existente en base de datos para evitar registros duplicados.
            if force_overwrite or (target_date and date_str == target_date):
                cursor.execute("SELECT id FROM REPORTES WHERE fecha = ?;", (date_str,))
                rep_row = cursor.fetchone()
                if rep_row:
                    rep_id = rep_row[0]
                    cursor.execute("DELETE FROM REGISTRO_PERSONAL WHERE id_reporte = ?;", (rep_id,))
                    cursor.execute("DELETE FROM REPORTES WHERE id = ?;", (rep_id,))
                    # Remover de existing_dates para que el script proceda a insertarlo de nuevo
                    existing_dates.discard(date_str)

            if date_str in existing_dates:
                continue
                
            print(f"Importando reporte de fecha {date_str} a la base de datos de Neon (modo optimizado)...")
            
            # Insertar reporte
            cursor.execute("INSERT INTO REPORTES (fecha, archivo) VALUES (?, ?);", (date_str, filename))
            cursor.execute("SELECT id FROM REPORTES WHERE fecha = ?;", (date_str,))
            report_id = cursor.fetchone()[0]
            
            records_to_insert = []
            
            for rec_id, record in records.items():
                cedula = record.get("CEDULA")
                nombre = record.get("APELLIDOS Y NOMBRES")
                subnovedad = record.get("SUBNOVEDAD")
                descripcion = record.get("DESCRIPCION", "")
                desde_val = record.get("DESDE", "")
                hasta_val = record.get("HASTA", "")
                
                if cedula is None or nombre is None:
                    continue
                try:
                    cedula_int = int(float(str(cedula).strip()))
                except ValueError:
                    continue
                if cedula_int <= 0:
                    continue
                    
                nombre = str(nombre).strip().upper()
                subnovedad = str(subnovedad).strip().upper() if subnovedad else "SIN NOVEDAD"
                
                # Obtener o crear Personal (Caché local primero)
                if cedula_int in personal_cache:
                    personal_id = personal_cache[cedula_int]
                else:
                    cursor.execute("INSERT INTO PERSONAL (cedula, nombre) VALUES (?, ?);", (cedula_int, nombre))
                    cursor.execute("SELECT id FROM PERSONAL WHERE cedula = ?;", (cedula_int,))
                    personal_id = cursor.fetchone()[0]
                    personal_cache[cedula_int] = personal_id
                    
                # Obtener o crear Subnovedad (Caché local primero)
                if subnovedad in subnovedades_cache:
                    subnovedad_id = subnovedades_cache[subnovedad]
                else:
                    cursor.execute("INSERT INTO SUB_NOVEDADES (nombre) VALUES (?);", (subnovedad,))
                    cursor.execute("SELECT id FROM SUB_NOVEDADES WHERE nombre = ?;", (subnovedad,))
                    subnovedad_id = cursor.fetchone()[0]
                    subnovedades_cache[subnovedad] = subnovedad_id
                    
                # Limpiar fechas
                fecha_inicio = None
                if desde_val:
                    try:
                        fecha_inicio = str(desde_val).split()[0]
                    except Exception:
                        pass
                fecha_final = None
                if hasta_val:
                    try:
                        fecha_final = str(hasta_val).split()[0]
                    except Exception:
                        pass
                        
                # Acumular registro para inserción por lote
                records_to_insert.append((
                    report_id,
                    personal_id,
                    subnovedad_id,
                    descripcion,
                    fecha_inicio,
                    fecha_final
                ))
            
            # Inserción en lote en PostgreSQL para evitar roundtrips de red N+1
            if records_to_insert:
                import psycopg2.extras
                psycopg2.extras.execute_values(
                    cursor._cursor,
                    """
                    INSERT INTO REGISTRO_PERSONAL 
                    (id_reporte, id_personal, id_sub_novedad, descripcion, fecha_inicio, fecha_final)
                    VALUES %s;
                    """,
                    records_to_insert
                )
                print(f"  -> {len(records_to_insert)} registros de personal cargados en lote con éxito.")
                
        db.commit()

@router.post("/sincronizar/drive")
def sincronizar_desde_drive(
    req: DriveSyncRequest,
    current_user: dict = Depends(get_current_user)
):
    """
    Ejecuta el script leer_carpetas.py y leer_archivos_excel.py 
    para descargar y procesar nuevos archivos desde Google Drive,
    e inserta los nuevos datos en la base de datos de Neon.
    """
    try:
        # 1. Obtener filtros de la consulta
        target_month = req.mes
        target_date = req.fecha
        target_dates = req.fechas  # Lista de fechas para modo multi-día
        
        # Si es tipo día, extraemos el mes para optimizar
        if req.tipo == "dia" and target_date:
            try:
                m_num = int(target_date.split("-")[1])
                reverse_map = {
                    1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
                    5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
                    9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
                }
                target_month = reverse_map.get(m_num)
            except Exception:
                pass
        elif req.tipo == "dias" and target_dates:
            # Modo multi-día: determinamos los meses únicos involucrados
            # y limpiamos target_date/target_month ya que usamos target_dates
            target_date = None
            target_month = None
        elif req.tipo == "todo":
            target_month = None
            target_date = None
            target_dates = None

        # 2. Ejecutar la lógica de leer_carpetas para actualizar listado_meses.json
        # Para modo multi-día, determinamos los meses necesarios y listamos cada uno
        from leer_carpetas import listar_dias_mes
        if req.tipo == "dias" and target_dates:
            # Listar carpetas para cada mes involucrado en las fechas seleccionadas
            reverse_map_local = {
                1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
                5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
                9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
            }
            meses_needed = set()
            for td in target_dates:
                try:
                    m_num = int(td.split("-")[1])
                    mes_name = reverse_map_local.get(m_num)
                    if mes_name:
                        meses_needed.add(mes_name)
                except Exception:
                    pass
            meses_nuevos = {}
            for mes_name in meses_needed:
                parcial = listar_dias_mes(target_month=mes_name)
                meses_nuevos.update(parcial)
        else:
            meses_nuevos = listar_dias_mes(target_month=target_month)
        
        # Leer listado_meses.json existente para no borrar información de otros meses
        existing_meses = {}
        try:
            with open("listado_meses.json", "r", encoding="utf-8") as ls:
                existing_meses = json.load(ls)
        except Exception:
            existing_meses = {}
            
        existing_meses.update(meses_nuevos)
        
        with open("listado_meses.json", "w", encoding="utf-8") as ls:
            json.dump(existing_meses, ls, indent=4, ensure_ascii=False)

        # 3. Ejecutar la lógica de leer_archivos_excel para descargar y actualizar las hojas mensuales
        # Pasamos db=None ya que obtener_hojas abre su propia conexion temporal rapida y la cierra de inmediato
        from leer_archivos_excel import obtener_hojas
        errors, sync_log = obtener_hojas(
            db=None,
            target_month=target_month,
            target_date=target_date,
            target_dates=target_dates,
            force_overwrite=req.overwrite
        )

        # 4. Sincronizar todos los JSONs locales nuevos a la base de datos (PostgreSQL Neon)
        # Para evitar InterfaceError: cursor already closed debido a timeouts idle de Neon en descargas largas,
        # abrimos una conexion dedicada y fresca de corta duracion únicamente para la insercion.
        from app.database import ConnectionWrapper
        import psycopg2
        
        print("Abriendo conexion dedicada de escritura a Neon...")
        raw_conn = psycopg2.connect(
            dbname="neondb",
            user="neondb_owner",
            password="npg_pPVueS4skO8j",
            host="ep-snowy-glade-aty6j16z-pooler.c-9.us-east-1.aws.neon.tech",
            sslmode="require"
        )
        db_write = ConnectionWrapper(raw_conn)
        try:
            sync_local_jsons_to_db(
                db=db_write,
                target_month=target_month,
                target_date=target_date,
                target_dates=target_dates,
                force_overwrite=req.overwrite
            )
        finally:
            db_write.close()
            print("Conexion dedicada cerrada exitosamente.")

        return {
            "status": "success",
            "message": "Sincronización con Google Drive completada exitosamente. Se actualizaron los datos en la nube.",
            "errors": errors,
            "logs": sync_log
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=500, 
            detail=f"Error al sincronizar con Google Drive: {str(e)}"
        )

