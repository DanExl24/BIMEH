import os
import json
import sqlite3
import psycopg2
import psycopg2.extras
import traceback
import csv
import io

class CursorWrapper:
    def __init__(self, cursor):
        self._cursor = cursor
        
    def execute(self, query, vars=None):
        if isinstance(query, str):
            query = query.replace("strftime('%m', fecha)", "to_char(to_date(fecha, 'YYYY-MM-DD'), 'MM')")
            query = query.replace("strftime('%d', fecha)", "to_char(to_date(fecha, 'YYYY-MM-DD'), 'DD')")
            query = query.replace("strftime('%m', r.fecha)", "to_char(to_date(r.fecha, 'YYYY-MM-DD'), 'MM')")
            query = query.replace("strftime('%d', r.fecha)", "to_char(to_date(r.fecha, 'YYYY-MM-DD'), 'DD')")
            query = query.replace('?', '%s')
        return self._cursor.execute(query, vars)
        
    def fetchone(self):
        return self._cursor.fetchone()
        
    def fetchall(self):
        return self._cursor.fetchall()
        
    def __getattr__(self, name):
        return getattr(self._cursor, name)

class ConnectionWrapper:
    def __init__(self, conn):
        self._conn = conn
        self._conn.cursor_factory = psycopg2.extras.DictCursor
        
    def cursor(self, *args, **kwargs):
        cursor = self._conn.cursor(*args, **kwargs)
        return CursorWrapper(cursor)
        
    def commit(self):
        return self._conn.commit()
        
    def rollback(self):
        return self._conn.rollback()
        
    def close(self):
        return self._conn.close()
        
    def execute(self, query, vars=None):
        if "PRAGMA" in query:
            return None
        if isinstance(query, str):
            query = query.replace("strftime('%m', fecha)", "to_char(to_date(fecha, 'YYYY-MM-DD'), 'MM')")
            query = query.replace("strftime('%d', fecha)", "to_char(to_date(fecha, 'YYYY-MM-DD'), 'DD')")
            query = query.replace("strftime('%m', r.fecha)", "to_char(to_date(r.fecha, 'YYYY-MM-DD'), 'MM')")
            query = query.replace("strftime('%d', r.fecha)", "to_char(to_date(r.fecha, 'YYYY-MM-DD'), 'DD')")
            query = query.replace('?', '%s')
        cursor = self.cursor()
        cursor.execute(query, vars)
        return cursor

    def __getattr__(self, name):
        return getattr(self._conn, name)
from datetime import datetime, timedelta
from typing import Optional, List, Dict, Any

from fastapi import FastAPI, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse, Response
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

# For Excel generation
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# For PDF generation
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

app = FastAPI(title="BIMEJ12 - Sistema de Reportes de Personal", version="1.0.0")

# Enable CORS for frontend development
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], # In production, restrict to frontend URL
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.middleware("http")
async def log_requests(request, call_next):
    start_time = datetime.now()
    method = request.method
    path = request.url.path
    print(f"--> [REQ] {method} {path}")
    try:
        response = await call_next(request)
        process_time = (datetime.now() - start_time).total_seconds() * 1000
        print(f"<-- [RES] {method} {path} status={response.status_code} ({process_time:.2f}ms)")
        return response
    except Exception as e:
        process_time = (datetime.now() - start_time).total_seconds() * 1000
        print(f"❌ [ERR] {method} {path} ({process_time:.2f}ms): {e}")
        traceback.print_exc()
        raise e

DATABASE_NAME = "bimeh"
DISPONIBLE_STATUSES = ["CDO UNIDAD", "AREA OPERACIONES"]

def get_db():
    raw_conn = psycopg2.connect(
        dbname="bimeh",
        user="postgres",
        password="postgres",
        host="localhost",
        port=5432
    )
    conn = ConnectionWrapper(raw_conn)
    try:
        yield conn
    finally:
        conn.close()

# Helpers
def get_month_dates(month_name: str) -> List[str]:
    month_order = {
        "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4,
        "MAYO": 5, "JUNIO": 6, "JULIO": 7, "AGOSTO": 8,
        "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12
    }
    month_num = month_order.get(month_name.upper(), 1)
    
    # Query database for available dates in this month
    raw_conn = psycopg2.connect(
        dbname="bimeh",
        user="postgres",
        password="postgres",
        host="localhost",
        port=5432
    )
    conn = ConnectionWrapper(raw_conn)
    cursor = conn.cursor()
    cursor.execute("SELECT fecha FROM REPORTES ORDER BY fecha ASC;")
    all_dates = [row[0] for row in cursor.fetchall()]
    conn.close()
    
    # Filter dates belonging to the target month (e.g. "2026-01-...")
    filtered = []
    for d in all_dates:
        try:
            dt = datetime.strptime(d, "%Y-%m-%d")
            if dt.month == month_num:
                filtered.append(d)
        except ValueError:
            continue
    return filtered

def parse_date(date_str: str) -> datetime:
    return datetime.strptime(date_str, "%Y-%m-%d")

# Pydantic models for request/response bodies (optional, query params are mostly used)
class KPIData(BaseModel):
    fecha: str
    total_personal: int
    disponibles: int
    novedades: int
    disponibilidad: float
    cambios_vs_ayer: int

# --- API ENDPOINTS ---

@app.get("/api/fechas")
def get_available_dates(db: sqlite3.Connection = Depends(get_db)):
    """Returns all dates that have reports, sorted chronologically."""
    cursor = db.cursor()
    cursor.execute("SELECT fecha FROM REPORTES ORDER BY fecha ASC;")
    dates = [row[0] for row in cursor.fetchall()]
    return dates

def get_report_ids_for_filter(db: sqlite3.Connection, mes: Optional[str] = None, dia: Optional[str] = None, fecha: Optional[str] = None) -> List[tuple]:
    cursor = db.cursor()
    
    if fecha:
        cursor.execute("SELECT id, fecha FROM REPORTES WHERE fecha = ?;", (fecha,))
        return cursor.fetchall()
        
    query = "SELECT id, fecha FROM REPORTES"
    where_clauses = []
    params = []
    
    month_order = {
        "ENERO": "01", "FEBRERO": "02", "MARZO": "03", "ABRIL": "04",
        "MAYO": "05", "JUNIO": "06", "JULIO": "07", "AGOSTO": "08",
        "SEPTIEMBRE": "09", "OCTUBRE": "10", "NOVIEMBRE": "11", "DICIEMBRE": "12"
    }
    
    if mes and mes.upper() in month_order:
        month_num = month_order[mes.upper()]
        where_clauses.append("strftime('%m', fecha) = ?")
        params.append(month_num)
        
    if dia:
        day_str = dia.zfill(2)
        where_clauses.append("strftime('%d', fecha) = ?")
        params.append(day_str)
        
    if where_clauses:
        query += " WHERE " + " AND ".join(where_clauses)
        
    query += " ORDER BY fecha ASC;"
    
    cursor.execute(query, params)
    return cursor.fetchall()

@app.get("/api/dashboard/kpis", response_model=KPIData)
def get_kpis(
    fecha: Optional[str] = Query(None, description="Fecha en formato YYYY-MM-DD"),
    mes: Optional[str] = Query(None, description="Nombre del mes (ENERO, FEBRERO...)"),
    dia: Optional[str] = Query(None, description="Día del mes (01, 02...)"),
    db: sqlite3.Connection = Depends(get_db)
):
    cursor = db.cursor()
    
    # If no parameters provided, default to the latest report date in DB
    if fecha is None and mes is None and dia is None:
        cursor.execute("SELECT fecha FROM REPORTES ORDER BY fecha DESC LIMIT 1;")
        latest_row = cursor.fetchone()
        if not latest_row:
            raise HTTPException(status_code=404, detail="No hay reportes cargados en el sistema.")
        fecha = latest_row[0]
        
    reports = get_report_ids_for_filter(db, mes, dia, fecha)
    if not reports:
        raise HTTPException(status_code=404, detail="No se encontraron reportes para el filtro seleccionado.")
        
    placeholders = ",".join("?" for _ in DISPONIBLE_STATUSES)
    cursor.execute(f"SELECT id FROM SUB_NOVEDADES WHERE nombre IN ({placeholders});", DISPONIBLE_STATUSES)
    available_ids = [row[0] for row in cursor.fetchall()]
    
    total_reports = len(reports)
    report_ids = [r[0] for r in reports]
    rep_placeholders = ",".join("?" for _ in report_ids)
    
    # Sum total personal
    cursor.execute(f"SELECT COUNT(*) FROM REGISTRO_PERSONAL WHERE id_reporte IN ({rep_placeholders});", report_ids)
    sum_total_personal = cursor.fetchone()[0]
    
    # Sum disponibles
    sum_disponibles = 0
    if available_ids:
        av_placeholders = ",".join("?" for _ in available_ids)
        cursor.execute(f"""
            SELECT COUNT(*) FROM REGISTRO_PERSONAL 
            WHERE id_reporte IN ({rep_placeholders}) 
            AND id_sub_novedad IN ({av_placeholders});
        """, (*report_ids, *available_ids))
        sum_disponibles = cursor.fetchone()[0]
        
    # Averages
    avg_total_personal = int(round(sum_total_personal / total_reports)) if total_reports > 0 else 0
    avg_disponibles = int(round(sum_disponibles / total_reports)) if total_reports > 0 else 0
    avg_novedades = avg_total_personal - avg_disponibles
    avg_disponibilidad = round((avg_disponibles / avg_total_personal * 100), 1) if avg_total_personal > 0 else 0.0
    
    # Changes vs yesterday summed
    total_cambios = 0
    for r_id, r_fecha in reports:
        cursor.execute("SELECT id FROM REPORTES WHERE fecha < ? ORDER BY fecha DESC LIMIT 1;", (r_fecha,))
        prev_row = cursor.fetchone()
        if prev_row:
            prev_report_id = prev_row[0]
            
            cursor.execute("SELECT id_personal, id_sub_novedad FROM REGISTRO_PERSONAL WHERE id_reporte = ?;", (r_id,))
            today_statuses = {row[0]: row[1] for row in cursor.fetchall()}
            
            cursor.execute("SELECT id_personal, id_sub_novedad FROM REGISTRO_PERSONAL WHERE id_reporte = ?;", (prev_report_id,))
            yesterday_statuses = {row[0]: row[1] for row in cursor.fetchall()}
            
            all_person_ids = set(today_statuses.keys()).union(yesterday_statuses.keys())
            for pid in all_person_ids:
                if today_statuses.get(pid) != yesterday_statuses.get(pid):
                    total_cambios += 1
                    
    # Generate custom fecha label
    if fecha:
        fecha_label = fecha
    elif mes and dia:
        fecha_label = f"{mes} - Día {dia}"
    elif mes:
        fecha_label = f"Mes: {mes}"
    elif dia:
        fecha_label = f"Día {dia} (Anual)"
    else:
        fecha_label = "Anual (Todo el año)"
        
    return KPIData(
        fecha=fecha_label,
        total_personal=avg_total_personal,
        disponibles=avg_disponibles,
        novedades=avg_novedades,
        disponibilidad=avg_disponibilidad,
        cambios_vs_ayer=total_cambios
    )

@app.get("/api/dashboard/evolucion")
def get_evolucion(
    mes: Optional[str] = Query(None, description="Nombre del mes (ENERO, FEBRERO...)"),
    dia: Optional[str] = Query(None, description="Día del mes (01, 02...)"),
    db: sqlite3.Connection = Depends(get_db)
):
    cursor = db.cursor()
    
    reports = get_report_ids_for_filter(db, mes, dia, None)
    
    placeholders = ",".join("?" for _ in DISPONIBLE_STATUSES)
    cursor.execute(f"SELECT id FROM SUB_NOVEDADES WHERE nombre IN ({placeholders});", DISPONIBLE_STATUSES)
    available_ids = [row[0] for row in cursor.fetchall()]
    
    evolution_data = []
    
    for r in reports:
        report_id, date_str = r[0], r[1]
        
        cursor.execute("SELECT COUNT(*) FROM REGISTRO_PERSONAL WHERE id_reporte = ?;", (report_id,))
        total = cursor.fetchone()[0]
        
        if available_ids:
            av_placeholders = ",".join("?" for _ in available_ids)
            cursor.execute(f"SELECT COUNT(*) FROM REGISTRO_PERSONAL WHERE id_reporte = ? AND id_sub_novedad IN ({av_placeholders});", (report_id, *available_ids))
            disponibles = cursor.fetchone()[0]
        else:
            disponibles = 0
            
        novedades = total - disponibles
        disponibilidad = round((disponibles / total * 100), 1) if total > 0 else 0.0
        
        evolution_data.append({
            "fecha": date_str,
            "total_personal": total,
            "disponibles": disponibles,
            "novedades": novedades,
            "disponibilidad": disponibilidad
        })
        
    return evolution_data

@app.get("/api/dashboard/novedades-frecuentes")
def get_novedades_frecuentes(
    fecha: Optional[str] = Query(None),
    mes: Optional[str] = Query(None),
    dia: Optional[str] = Query(None),
    db: sqlite3.Connection = Depends(get_db)
):
    cursor = db.cursor()
    
    if fecha is None and mes is None and dia is None:
        cursor.execute("SELECT fecha FROM REPORTES ORDER BY fecha DESC LIMIT 1;")
        latest_row = cursor.fetchone()
        if latest_row:
            fecha = latest_row[0]
            
    reports = get_report_ids_for_filter(db, mes, dia, fecha)
    if not reports:
        return []
        
    report_ids = [r[0] for r in reports]
    rep_placeholders = ",".join("?" for _ in report_ids)
    
    ex_placeholders = ",".join("?" for _ in DISPONIBLE_STATUSES)
    
    cursor.execute(f"""
        SELECT sn.nombre, COUNT(*) as cantidad 
        FROM REGISTRO_PERSONAL rp 
        JOIN SUB_NOVEDADES sn ON rp.id_sub_novedad = sn.id
        WHERE rp.id_reporte IN ({rep_placeholders})
        AND sn.nombre NOT IN ({ex_placeholders})
        GROUP BY sn.nombre 
        ORDER BY cantidad DESC;
    """, (*report_ids, *DISPONIBLE_STATUSES))
    
    rows = cursor.fetchall()
    return [{"novedad": row[0], "cantidad": row[1]} for row in rows]

@app.get("/api/dashboard/distribucion")
def get_distribucion(
    fecha: Optional[str] = Query(None),
    mes: Optional[str] = Query(None),
    dia: Optional[str] = Query(None),
    db: sqlite3.Connection = Depends(get_db)
):
    cursor = db.cursor()
    
    if fecha is None and mes is None and dia is None:
        cursor.execute("SELECT fecha FROM REPORTES ORDER BY fecha DESC LIMIT 1;")
        latest_row = cursor.fetchone()
        if latest_row:
            fecha = latest_row[0]
            
    reports = get_report_ids_for_filter(db, mes, dia, fecha)
    if not reports:
        return []
        
    report_ids = [r[0] for r in reports]
    rep_placeholders = ",".join("?" for _ in report_ids)
    
    cursor.execute(f"""
        SELECT sn.nombre, COUNT(*) as cantidad
        FROM REGISTRO_PERSONAL rp
        JOIN SUB_NOVEDADES sn ON rp.id_sub_novedad = sn.id
        WHERE rp.id_reporte IN ({rep_placeholders})
        GROUP BY sn.nombre
        ORDER BY cantidad DESC;
    """, report_ids)
    
    rows = cursor.fetchall()
    total = sum(row[1] for row in rows)
    num_days = len(reports)
    
    dist = []
    for row in rows:
        nombre = row[0]
        cantidad_total = row[1]
        cantidad_avg = int(round(cantidad_total / num_days)) if num_days > 0 else 0
        pct = round((cantidad_total / total * 100), 1) if total > 0 else 0.0
        
        categoria = "DISPONIBLE" if nombre in DISPONIBLE_STATUSES else "NOVEDAD"
        
        dist.append({
            "subnovedad": nombre,
            "cantidad": cantidad_avg,
            "porcentaje": pct,
            "categoria": categoria
        })
        
    return dist

@app.get("/api/dashboard/cambios")
def get_cambios(
    fecha: Optional[str] = Query(None),
    mes: Optional[str] = Query(None),
    dia: Optional[str] = Query(None),
    db: sqlite3.Connection = Depends(get_db)
):
    cursor = db.cursor()
    
    if fecha is None and mes is None and dia is None:
        cursor.execute("SELECT fecha FROM REPORTES ORDER BY fecha DESC LIMIT 1;")
        latest_row = cursor.fetchone()
        if latest_row:
            fecha = latest_row[0]
            
    reports = get_report_ids_for_filter(db, mes, dia, fecha)
    if not reports:
        return {"entraron_novedades": [], "volvieron_disponibles": [], "otros_cambios": []}
        
    entraron_novedades = []
    volvieron_disponibles = []
    otros_cambios = []
    
    seen_changes = set()
    
    for r_id, r_fecha in reports:
        cursor.execute("SELECT id FROM REPORTES WHERE fecha < ? ORDER BY fecha DESC LIMIT 1;", (r_fecha,))
        prev_row = cursor.fetchone()
        if not prev_row:
            continue
        prev_report_id = prev_row[0]
        
        cursor.execute("""
            SELECT rp.id_personal, p.cedula, p.nombre, sn.nombre
            FROM REGISTRO_PERSONAL rp
            JOIN PERSONAL p ON rp.id_personal = p.id
            JOIN SUB_NOVEDADES sn ON rp.id_sub_novedad = sn.id
            WHERE rp.id_reporte = ?;
        """, (r_id,))
        today_map = {row[0]: {"cedula": row[1], "nombre": row[2], "subnovedad": row[3]} for row in cursor.fetchall()}
        
        cursor.execute("""
            SELECT rp.id_personal, p.cedula, p.nombre, sn.nombre
            FROM REGISTRO_PERSONAL rp
            JOIN PERSONAL p ON rp.id_personal = p.id
            JOIN SUB_NOVEDADES sn ON rp.id_sub_novedad = sn.id
            WHERE rp.id_reporte = ?;
        """, (prev_report_id,))
        prev_map = {row[0]: {"cedula": row[1], "nombre": row[2], "subnovedad": row[3]} for row in cursor.fetchall()}
        
        all_pids = set(today_map.keys()).union(prev_map.keys())
        
        for pid in all_pids:
            today_data = today_map.get(pid)
            prev_data = prev_map.get(pid)
            
            if today_data and prev_data:
                t_nov = today_data["subnovedad"]
                p_nov = prev_data["subnovedad"]
                
                if t_nov != p_nov:
                    change_key = (today_data["cedula"], p_nov, t_nov)
                    if change_key in seen_changes:
                        continue
                    seen_changes.add(change_key)
                    
                    change_item = {
                        "cedula": today_data["cedula"],
                        "nombre": today_data["nombre"],
                        "novedad_anterior": p_nov,
                        "novedad_nueva": t_nov,
                        "fecha": r_fecha
                    }
                    
                    p_disp = p_nov in DISPONIBLE_STATUSES
                    t_disp = t_nov in DISPONIBLE_STATUSES
                    
                    if p_disp and not t_disp:
                        entraron_novedades.append(change_item)
                    elif not p_disp and t_disp:
                        volvieron_disponibles.append(change_item)
                    else:
                        otros_cambios.append(change_item)
            elif today_data and not prev_data:
                t_nov = today_data["subnovedad"]
                change_key = (today_data["cedula"], "NO PRESENTADO", t_nov)
                if change_key in seen_changes:
                    continue
                seen_changes.add(change_key)
                
                change_item = {
                    "cedula": today_data["cedula"],
                    "nombre": today_data["nombre"],
                    "novedad_anterior": "NO PRESENTADO",
                    "novedad_nueva": t_nov,
                    "fecha": r_fecha
                }
                if t_nov in DISPONIBLE_STATUSES:
                    volvieron_disponibles.append(change_item)
                else:
                    entraron_novedades.append(change_item)
            elif prev_data and not today_data:
                p_nov = prev_data["subnovedad"]
                change_key = (prev_data["cedula"], p_nov, "RETIRADO / NO PRESENTADO")
                if change_key in seen_changes:
                    continue
                seen_changes.add(change_key)
                
                change_item = {
                    "cedula": prev_data["cedula"],
                    "nombre": prev_data["nombre"],
                    "novedad_anterior": p_nov,
                    "novedad_nueva": "RETIRADO / NO PRESENTADO",
                    "fecha": r_fecha
                }
                if p_nov in DISPONIBLE_STATUSES:
                    entraron_novedades.append(change_item)
                else:
                    otros_cambios.append(change_item)
                    
    return {
        "entraron_novedades": entraron_novedades[:150],
        "volvieron_disponibles": volvieron_disponibles[:150],
        "otros_cambios": otros_cambios[:150]
    }

@app.get("/api/personal/buscar")
def buscar_personal(q: str = Query(..., min_length=2), db: sqlite3.Connection = Depends(get_db)):
    cursor = db.cursor()
    
    # Try search by digit (cedula) or by name pattern
    search_pattern = f"%{q.upper()}%"
    cursor.execute("""
        SELECT cedula, nombre, CASE WHEN fecha_retiro IS NULL THEN 'ACTIVO' ELSE 'RETIRADO' END as estado, fecha_retiro 
        FROM PERSONAL 
        WHERE CAST(cedula AS TEXT) LIKE ? OR nombre LIKE ?
        LIMIT 50;
    """, (search_pattern, search_pattern))
    
    rows = cursor.fetchall()
    return [{
        "cedula": row[0],
        "nombre": row[1],
        "estado": row[2],
        "fecha_retiro": row[3]
    } for row in rows]

@app.get("/api/personal/{cedula}")
def get_personal_detalle(cedula: int, db: sqlite3.Connection = Depends(get_db)):
    cursor = db.cursor()
    
    # Find personnel
    cursor.execute("SELECT id, nombre, CASE WHEN fecha_retiro IS NULL THEN 'ACTIVO' ELSE 'RETIRADO' END as estado, fecha_retiro FROM PERSONAL WHERE cedula = ?;", (cedula,))
    p_row = cursor.fetchone()
    if not p_row:
        raise HTTPException(status_code=404, detail="Personal no encontrado.")
        
    p_id, nombre, estado, fecha_retiro = p_row[0], p_row[1], p_row[2], p_row[3]
    
    # Get all reports the user is in
    cursor.execute("""
        SELECT r.fecha, sn.nombre 
        FROM REGISTRO_PERSONAL rp
        JOIN REPORTES r ON rp.id_reporte = r.id
        JOIN SUB_NOVEDADES sn ON rp.id_sub_novedad = sn.id
        WHERE rp.id_personal = ?
        ORDER BY r.fecha ASC;
    """, (p_id,))
    
    records = cursor.fetchall()
    total_dias = len(records)
    
    if total_dias == 0:
        return {
            "cedula": cedula,
            "nombre": nombre,
            "estado": estado,
            "fecha_retiro": fecha_retiro,
            "primer_registro_fecha": None,
            "ultimo_registro_fecha": None,
            "total_dias": 0,
            "tiempo_disponible_pct": 0,
            "tiempo_novedades_pct": 0,
            "total_novedades": 0,
            "promedio_duracion_novedades": 0.0,
            "ultima_novedad": None
        }
        
    primer_registro_fecha = records[0][0]
    ultimo_registro_fecha = records[-1][0]
    
    disponibles_dias = sum(1 for r in records if r[1] in DISPONIBLE_STATUSES)
    tiempo_disponible_pct = round((disponibles_dias / total_dias * 100), 1)
    tiempo_novedades_pct = round((100.0 - tiempo_disponible_pct), 1)
    total_novedades = total_dias - disponibles_dias
    
    # Calculate average novelty length
    # An event is a run of consecutive reports in a non-available status
    novelty_runs = []
    current_run = 0
    
    for r in records:
        is_available = r[1] in DISPONIBLE_STATUSES
        if not is_available:
            current_run += 1
        else:
            if current_run > 0:
                novelty_runs.append(current_run)
                current_run = 0
                
    if current_run > 0:
        novelty_runs.append(current_run)
        
    promedio_duracion_novedades = round(sum(novelty_runs) / len(novelty_runs), 1) if novelty_runs else 0.0
    
    ultima_novedad = records[-1][1] if records else None
    
    return {
        "cedula": cedula,
        "nombre": nombre,
        "estado": estado,
        "fecha_retiro": fecha_retiro,
        "primer_registro_fecha": primer_registro_fecha,
        "ultimo_registro_fecha": ultimo_registro_fecha,
        "total_dias": total_dias,
        "tiempo_disponible_pct": tiempo_disponible_pct,
        "tiempo_novedades_pct": tiempo_novedades_pct,
        "total_novedades": total_novedades,
        "promedio_duracion_novedades": promedio_duracion_novedades,
        "ultima_novedad": ultima_novedad
    }

@app.get("/api/personal/{cedula}/historial")
def get_personal_historial(cedula: int, db: sqlite3.Connection = Depends(get_db)):
    cursor = db.cursor()
    
    cursor.execute("SELECT id FROM PERSONAL WHERE cedula = ?;", (cedula,))
    p_row = cursor.fetchone()
    if not p_row:
        raise HTTPException(status_code=404, detail="Personal no encontrado.")
    p_id = p_row[0]
    
    cursor.execute("""
        SELECT r.fecha, sn.nombre, rp.descripcion, rp.fecha_inicio, rp.fecha_final
        FROM REGISTRO_PERSONAL rp
        JOIN REPORTES r ON rp.id_reporte = r.id
        JOIN SUB_NOVEDADES sn ON rp.id_sub_novedad = sn.id
        WHERE rp.id_personal = ?
        ORDER BY r.fecha DESC;
    """, (p_id,))
    
    rows = cursor.fetchall()
    return [{
        "fecha": row[0],
        "subnovedad": row[1],
        "descripcion": row[2],
        "desde": row[3],
        "hasta": row[4]
    } for row in rows]

@app.get("/api/personal/{cedula}/acumulado")
def get_personal_acumulado(cedula: int, db: sqlite3.Connection = Depends(get_db)):
    cursor = db.cursor()
    
    cursor.execute("SELECT id FROM PERSONAL WHERE cedula = ?;", (cedula,))
    p_row = cursor.fetchone()
    if not p_row:
        raise HTTPException(status_code=404, detail="Personal no encontrado.")
    p_id = p_row[0]
    
    cursor.execute("""
        SELECT sn.nombre, COUNT(*) as dias
        FROM REGISTRO_PERSONAL rp
        JOIN SUB_NOVEDADES sn ON rp.id_sub_novedad = sn.id
        WHERE rp.id_personal = ?
        GROUP BY sn.nombre
        ORDER BY dias DESC;
    """, (p_id,))
    
    rows = cursor.fetchall()
    return [{"subnovedad": row[0], "dias": row[1]} for row in rows]

@app.get("/api/reportes/dia")
def get_reporte_dia(fecha: str = Query(...), db: sqlite3.Connection = Depends(get_db)):
    cursor = db.cursor()
    
    cursor.execute("SELECT id FROM REPORTES WHERE fecha = ?;", (fecha,))
    r_row = cursor.fetchone()
    if not r_row:
        return []
    r_id = r_row[0]
    
    cursor.execute("""
        SELECT p.cedula, p.nombre, sn.nombre as subnovedad, rp.descripcion, rp.fecha_inicio, rp.fecha_final
        FROM REGISTRO_PERSONAL rp
        JOIN PERSONAL p ON rp.id_personal = p.id
        JOIN SUB_NOVEDADES sn ON rp.id_sub_novedad = sn.id
        WHERE rp.id_reporte = ?
        ORDER BY p.nombre ASC;
    """, (r_id,))
    
    rows = cursor.fetchall()
    return [{
        "cedula": row[0],
        "nombre": row[1],
        "subnovedad": row[2],
        "descripcion": row[3],
        "desde": row[4],
        "hasta": row[5]
    } for row in rows]

@app.get("/api/reportes/calendario")
def get_calendario(mes: str = Query(...), db: sqlite3.Connection = Depends(get_db)):
    dates = get_month_dates(mes)
    if not dates:
        return []
        
    cursor = db.cursor()
    
    placeholders = ",".join("?" for _ in DISPONIBLE_STATUSES)
    cursor.execute(f"SELECT id FROM SUB_NOVEDADES WHERE nombre IN ({placeholders});", DISPONIBLE_STATUSES)
    avail_ids = [row[0] for row in cursor.fetchall()]
    
    avail_placeholders = ",".join("?" for _ in avail_ids)
    
    cal_data = []
    
    for d in dates:
        cursor.execute("SELECT id FROM REPORTES WHERE fecha = ?;", (d,))
        r_id = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM REGISTRO_PERSONAL WHERE id_reporte = ?;", (r_id,))
        total = cursor.fetchone()[0]
        
        if total > 0 and avail_ids:
            cursor.execute(f"SELECT COUNT(*) FROM REGISTRO_PERSONAL WHERE id_reporte = ? AND id_sub_novedad IN ({avail_placeholders});", (r_id, *avail_ids))
            disponibles = cursor.fetchone()[0]
        else:
            disponibles = 0
            
        novedades = total - disponibles
        pct = round((disponibles / total * 100), 1) if total > 0 else 0.0
        
        cal_data.append({
            "fecha": d,
            "disponibilidad": pct,
            "total_personal": total,
            "disponibles": disponibles,
            "novedades": novedades
        })
        
    return cal_data

@app.get("/api/stats/ranking")
def get_stats_rankings(db: sqlite3.Connection = Depends(get_db)):
    cursor = db.cursor()
    
    # Most frequent subnovedades globally
    cursor.execute("""
        SELECT sn.nombre, COUNT(*) as total_dias
        FROM REGISTRO_PERSONAL rp
        JOIN SUB_NOVEDADES sn ON rp.id_sub_novedad = sn.id
        GROUP BY sn.nombre
        ORDER BY total_dias DESC;
    """)
    global_rank = [{"subnovedad": r[0], "dias_acumulados": r[1]} for r in cursor.fetchall()]
    
    # People with most novelty days
    placeholders = ",".join("?" for _ in DISPONIBLE_STATUSES)
    cursor.execute(f"""
        SELECT p.cedula, p.nombre, COUNT(*) as dias_novedad
        FROM REGISTRO_PERSONAL rp
        JOIN PERSONAL p ON rp.id_personal = p.id
        JOIN SUB_NOVEDADES sn ON rp.id_sub_novedad = sn.id
        WHERE sn.nombre NOT IN ({placeholders})
        GROUP BY p.id, p.cedula, p.nombre
        ORDER BY dias_novedad DESC
        LIMIT 15;
    """, DISPONIBLE_STATUSES)
    most_novelties_people = [{"cedula": r[0], "nombre": r[1], "dias_novedad": r[2]} for r in cursor.fetchall()]
    
    return {
        "global_rank": global_rank,
        "most_novelties_people": most_novelties_people
    }

@app.get("/api/stats/heatmap")
def get_stats_heatmap(mes: str = Query(...), db: sqlite3.Connection = Depends(get_db)):
    dates = get_month_dates(mes)
    if not dates:
        return {"fechas": [], "data": []}
        
    cursor = db.cursor()
    
    # Get all reports in this month and map them to their index
    cursor.execute(f"SELECT id, fecha FROM REPORTES WHERE fecha IN ({','.join('?' for _ in dates)}) ORDER BY fecha ASC;", dates)
    reports = cursor.fetchall()
    report_ids = [r[0] for r in reports]
    report_id_to_date = {r[0]: r[1] for r in reports}
    
    # Find all personnel that have at least one record in these reports
    cursor.execute(f"""
        SELECT DISTINCT p.id, p.cedula, p.nombre
        FROM REGISTRO_PERSONAL rp
        JOIN PERSONAL p ON rp.id_personal = p.id
        WHERE rp.id_reporte IN ({','.join('?' for _ in report_ids)})
        ORDER BY p.nombre ASC;
    """, report_ids)
    personnel = cursor.fetchall()
    
    # Fetch all records for these people on these reports
    cursor.execute(f"""
        SELECT rp.id_personal, rp.id_reporte, sn.nombre
        FROM REGISTRO_PERSONAL rp
        JOIN SUB_NOVEDADES sn ON rp.id_sub_novedad = sn.id
        WHERE rp.id_reporte IN ({','.join('?' for _ in report_ids)});
    """, report_ids)
    
    records = cursor.fetchall()
    # Create lookup map: (person_id, date) -> subnovedad_name
    record_map = {}
    for r in records:
        pid, rid, sub_name = r[0], r[1], r[2]
        date_str = report_id_to_date.get(rid)
        if date_str:
            record_map[(pid, date_str)] = sub_name
            
    heatmap_data = []
    for p in personnel:
        pid, cedula, nombre = p[0], p[1], p[2]
        estados = []
        for d in dates:
            # If not in registry for that day, mark as NULL or NO PRESENTADO or RETIRADO
            est = record_map.get((pid, d))
            if est is None:
                # Check if person was retired before this date or not yet registered
                # Let's just output "NO PRESENTADO"
                est = "N/A"
            estados.append(est)
            
        heatmap_data.append({
            "cedula": cedula,
            "nombre": nombre,
            "estados": estados
        })
        
    return {
        "fechas": dates,
        "data": heatmap_data
    }

@app.get("/api/alertas/inconsistencias")
def get_inconsistencias():
    """Reads processing_report.json and returns log of issues."""
    try:
        with open("processing_report.json", "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return []

# --- EXPORT ROUTERS ---

@app.get("/api/exportar/csv")
def exportar_csv(
    tipo: str = Query(..., description="dia, mes, personal, personal_db, subnovedades, consolidado_mensual o historial_novedades"),
    fecha: Optional[str] = Query(None),
    mes: Optional[str] = Query(None),
    cedula: Optional[int] = Query(None),
    db: sqlite3.Connection = Depends(get_db)
):
    output = io.StringIO()
    writer = csv.writer(output)
    
    cursor = db.cursor()
    
    if tipo == "dia" and fecha:
        writer.writerow(["CEDULA", "APELLIDOS Y NOMBRES", "SUBNOVEDAD", "DESCRIPCION", "DESDE", "HASTA"])
        cursor.execute("SELECT id FROM REPORTES WHERE fecha = ?;", (fecha,))
        r_row = cursor.fetchone()
        if r_row:
            cursor.execute("""
                SELECT p.cedula, p.nombre, sn.nombre, rp.descripcion, rp.fecha_inicio, rp.fecha_final
                FROM REGISTRO_PERSONAL rp
                JOIN PERSONAL p ON rp.id_personal = p.id
                JOIN SUB_NOVEDADES sn ON rp.id_sub_novedad = sn.id
                WHERE rp.id_reporte = ?
                ORDER BY p.nombre ASC;
            """, (r_row[0],))
            for row in cursor.fetchall():
                writer.writerow(list(row))
        filename = f"reporte_dia_{fecha}.csv"
        
    elif tipo == "mes" and mes:
        writer.writerow(["FECHA", "TOTAL PERSONAL", "DISPONIBLES", "NOVEDADES", "DISPONIBILIDAD %"])
        dates = get_month_dates(mes)
        for d in dates:
            cursor.execute("SELECT id FROM REPORTES WHERE fecha = ?;", (d,))
            r_id = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) FROM REGISTRO_PERSONAL WHERE id_reporte = ?;", (r_id,))
            total = cursor.fetchone()[0]
            
            placeholders = ",".join("%s" for _ in DISPONIBLE_STATUSES)
            cursor.execute(f"""
                SELECT COUNT(*) FROM REGISTRO_PERSONAL 
                WHERE id_reporte = ? AND id_sub_novedad IN (
                    SELECT id FROM SUB_NOVEDADES WHERE nombre IN ({placeholders})
                );
            """, (r_id, *DISPONIBLE_STATUSES))
            disp = cursor.fetchone()[0]
            nov = total - disp
            pct = round((disp / total * 100), 1) if total > 0 else 0.0
            writer.writerow([d, total, disp, nov, pct])
        filename = f"reporte_mensual_{mes}.csv"
        
    elif tipo == "personal" and cedula:
        writer.writerow(["FECHA", "SUBNOVEDAD", "DESCRIPCION", "DESDE", "HASTA"])
        cursor.execute("SELECT id, nombre FROM PERSONAL WHERE cedula = ?;", (cedula,))
        p_row = cursor.fetchone()
        if p_row:
            cursor.execute("""
                SELECT r.fecha, sn.nombre, rp.descripcion, rp.fecha_inicio, rp.fecha_final
                FROM REGISTRO_PERSONAL rp
                JOIN REPORTES r ON rp.id_reporte = r.id
                JOIN SUB_NOVEDADES sn ON rp.id_sub_novedad = sn.id
                WHERE rp.id_personal = ?
                ORDER BY r.fecha DESC;
            """, (p_row[0],))
            for row in cursor.fetchall():
                writer.writerow(list(row))
        filename = f"historial_cedula_{cedula}.csv"
        
    elif tipo == "personal_db":
        writer.writerow(["CEDULA", "APELLIDOS Y NOMBRES", "ESTADO", "FECHA RETIRO"])
        cursor.execute("""
            SELECT cedula, nombre, CASE WHEN fecha_retiro IS NULL THEN 'ACTIVO' ELSE 'RETIRADO' END as estado, fecha_retiro
            FROM PERSONAL
            ORDER BY nombre ASC;
        """)
        for row in cursor.fetchall():
            writer.writerow(list(row))
        filename = "base_datos_personal.csv"
        
    elif tipo == "subnovedades":
        writer.writerow(["ID", "NOMBRE NOVEDAD"])
        cursor.execute("SELECT id, nombre FROM SUB_NOVEDADES ORDER BY nombre ASC;")
        for row in cursor.fetchall():
            writer.writerow(list(row))
        filename = "catalogo_subnovedades.csv"
        
    elif tipo == "consolidado_mensual" and mes:
        dates = get_month_dates(mes)
        if not dates:
            raise HTTPException(status_code=400, detail="No hay reportes para el mes especificado.")
        
        placeholders = ",".join("%s" for _ in dates)
        cursor.execute(f"SELECT id, fecha FROM REPORTES WHERE fecha IN ({placeholders}) ORDER BY fecha ASC;", dates)
        reports_db = cursor.fetchall()
        report_ids = [r[0] for r in reports_db]
        report_dates = [r[1] for r in reports_db]
        
        headers = ["CEDULA", "INTEGRANTE"] + [f"Dia {d.split('-')[2]}" for d in report_dates]
        writer.writerow(headers)
        
        if report_ids:
            rep_placeholders = ",".join("%s" for _ in report_ids)
            cursor.execute(f"""
                SELECT p.cedula, p.nombre, rp.id_reporte, sn.nombre as subnovedad
                FROM REGISTRO_PERSONAL rp
                JOIN PERSONAL p ON rp.id_personal = p.id
                JOIN SUB_NOVEDADES sn ON rp.id_sub_novedad = sn.id
                WHERE rp.id_reporte IN ({rep_placeholders})
                ORDER BY p.nombre ASC;
            """, report_ids)
            
            person_map = {}
            for row in cursor.fetchall():
                key = (row[0], row[1])
                if key not in person_map:
                    person_map[key] = {}
                person_map[key][row[2]] = row[3]
                
            for (cedula, nombre), reports_dict in sorted(person_map.items(), key=lambda x: x[0][1]):
                row_data = [cedula, nombre]
                for r_id in report_ids:
                    row_data.append(reports_dict.get(r_id, "N/A"))
                writer.writerow(row_data)
        filename = f"consolidado_mensual_{mes}.csv"
        
    elif tipo == "historial_novedades":
        writer.writerow(["CEDULA", "APELLIDOS Y NOMBRES", "SUBNOVEDAD", "DESCRIPCION", "DESDE", "HASTA", "FECHA REPORTE"])
        cursor.execute("""
            SELECT p.cedula, p.nombre, sn.nombre, rp.descripcion, rp.fecha_inicio, rp.fecha_final, r.fecha
            FROM REGISTRO_PERSONAL rp
            JOIN PERSONAL p ON rp.id_personal = p.id
            JOIN SUB_NOVEDADES sn ON rp.id_sub_novedad = sn.id
            JOIN REPORTES r ON rp.id_reporte = r.id
            ORDER BY r.fecha DESC, p.nombre ASC;
        """)
        for row in cursor.fetchall():
            writer.writerow(list(row))
        filename = "historial_completo_novedades.csv"
        
    else:
        raise HTTPException(status_code=400, detail="Parámetros inválidos para la exportación.")
        
    response = StreamingResponse(io.BytesIO(output.getvalue().encode("utf-8-sig")), media_type="text/csv")
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response

@app.get("/api/exportar/excel")
def exportar_excel(
    tipo: str = Query(...),
    fecha: Optional[str] = Query(None),
    mes: Optional[str] = Query(None),
    cedula: Optional[int] = Query(None),
    db: sqlite3.Connection = Depends(get_db)
):
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Styles
    title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    normal_font = Font(name="Calibri", size=11)
    
    title_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid") # Dark gray
    header_fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid") # Lighter dark gray
    
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    cursor = db.cursor()
    
    if tipo == "dia" and fecha:
        ws.title = f"Reporte {fecha}"
        
        # Title Row
        ws.merge_cells("A1:F1")
        ws["A1"] = f"BIMEJ12 — REPORTE DIARIO DE PERSONAL — {fecha}"
        ws["A1"].font = title_font
        ws["A1"].fill = title_fill
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40
        
        # Headers
        headers = ["CÉDULA", "APELLIDOS Y NOMBRES", "SUBNOVEDAD", "DESCRIPCIÓN", "DESDE", "HASTA"]
        ws.append([]) # Blank row
        ws.append(headers)
        
        # Style headers
        ws.row_dimensions[3].height = 25
        for col_idx in range(1, 7):
            cell = ws.cell(row=3, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        cursor.execute("SELECT id FROM REPORTES WHERE fecha = ?;", (fecha,))
        r_row = cursor.fetchone()
        if r_row:
            cursor.execute("""
                SELECT p.cedula, p.nombre, sn.nombre, rp.descripcion, rp.fecha_inicio, rp.fecha_final
                FROM REGISTRO_PERSONAL rp
                JOIN PERSONAL p ON rp.id_personal = p.id
                JOIN SUB_NOVEDADES sn ON rp.id_sub_novedad = sn.id
                WHERE rp.id_reporte = ?
                ORDER BY p.nombre ASC;
            """, (r_row[0],))
            for row in cursor.fetchall():
                ws.append(list(row))
                
        # Format data rows
        for r_idx in range(4, ws.max_row + 1):
            ws.row_dimensions[r_idx].height = 20
            for c_idx in range(1, 7):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = normal_font
                cell.border = thin_border
                if c_idx == 1:
                    cell.alignment = Alignment(horizontal="left")
                elif c_idx in (5, 6):
                    cell.alignment = Alignment(horizontal="center")
                    
        filename = f"reporte_diario_{fecha}.xlsx"
        
    elif tipo == "mes" and mes:
        ws.title = f"Resumen {mes}"
        
        ws.merge_cells("A1:E1")
        ws["A1"] = f"BIMEJ12 — RESUMEN MENSUAL — {mes.upper()}"
        ws["A1"].font = title_font
        ws["A1"].fill = title_fill
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40
        
        headers = ["FECHA", "TOTAL PERSONAL", "DISPONIBLES", "NOVEDADES", "DISPONIBILIDAD %"]
        ws.append([])
        ws.append(headers)
        
        ws.row_dimensions[3].height = 25
        for col_idx in range(1, 6):
            cell = ws.cell(row=3, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        dates = get_month_dates(mes)
        for d in dates:
            cursor.execute("SELECT id FROM REPORTES WHERE fecha = ?;", (d,))
            r_id = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) FROM REGISTRO_PERSONAL WHERE id_reporte = ?;", (r_id,))
            total = cursor.fetchone()[0]
            
            placeholders = ",".join("%s" for _ in DISPONIBLE_STATUSES)
            cursor.execute(f"""
                SELECT COUNT(*) FROM REGISTRO_PERSONAL 
                WHERE id_reporte = ? AND id_sub_novedad IN (
                    SELECT id FROM SUB_NOVEDADES WHERE nombre IN ({placeholders})
                );
            """, (r_id, *DISPONIBLE_STATUSES))
            disp = cursor.fetchone()[0]
            nov = total - disp
            pct = round((disp / total * 100), 1) if total > 0 else 0.0
            ws.append([d, total, disp, nov, pct])
            
        for r_idx in range(4, ws.max_row + 1):
            ws.row_dimensions[r_idx].height = 20
            for c_idx in range(1, 6):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = normal_font
                cell.border = thin_border
                if c_idx == 1:
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="right")
                    
        filename = f"reporte_mensual_{mes}.xlsx"
        
    elif tipo == "personal" and cedula:
        cursor.execute("SELECT id, nombre, CASE WHEN fecha_retiro IS NULL THEN 'ACTIVO' ELSE 'RETIRADO' END as estado FROM PERSONAL WHERE cedula = ?;", (cedula,))
        p_row = cursor.fetchone()
        nombre = p_row[1] if p_row else "Desconocido"
        estado = p_row[2] if p_row else "Desconocido"
        
        ws.title = "Historial"
        
        ws.merge_cells("A1:E1")
        ws["A1"] = f"HISTORIAL INDIVIDUAL: {nombre} ({cedula}) - {estado}"
        ws["A1"].font = title_font
        ws["A1"].fill = title_fill
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40
        
        headers = ["FECHA", "SUBNOVEDAD", "DESCRIPCIÓN", "DESDE", "HASTA"]
        ws.append([])
        ws.append(headers)
        
        ws.row_dimensions[3].height = 25
        for col_idx in range(1, 6):
            cell = ws.cell(row=3, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        if p_row:
            cursor.execute("""
                SELECT r.fecha, sn.nombre, rp.descripcion, rp.fecha_inicio, rp.fecha_final
                FROM REGISTRO_PERSONAL rp
                JOIN REPORTES r ON rp.id_reporte = r.id
                JOIN SUB_NOVEDADES sn ON rp.id_sub_novedad = sn.id
                WHERE rp.id_personal = ?
                ORDER BY r.fecha DESC;
            """, (p_row[0],))
            for row in cursor.fetchall():
                ws.append(list(row))
                
        for r_idx in range(4, ws.max_row + 1):
            ws.row_dimensions[r_idx].height = 20
            for c_idx in range(1, 6):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = normal_font
                cell.border = thin_border
                if c_idx == 1:
                    cell.alignment = Alignment(horizontal="center")
                elif c_idx in (4, 5):
                    cell.alignment = Alignment(horizontal="center")
                    
        filename = f"historial_personal_{cedula}.xlsx"
        
    elif tipo == "personal_db":
        ws.title = "Base Personal"
        ws.merge_cells("A1:D1")
        ws["A1"] = "BIMEJ12 — BASE DE DATOS DE PERSONAL"
        ws["A1"].font = title_font
        ws["A1"].fill = title_fill
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40
        
        headers = ["CÉDULA", "APELLIDOS Y NOMBRES", "ESTADO", "FECHA RETIRO"]
        ws.append([])
        ws.append(headers)
        ws.row_dimensions[3].height = 25
        
        for col_idx in range(1, 5):
            cell = ws.cell(row=3, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        cursor.execute("""
            SELECT cedula, nombre, CASE WHEN fecha_retiro IS NULL THEN 'ACTIVO' ELSE 'RETIRADO' END as estado, fecha_retiro
            FROM PERSONAL
            ORDER BY nombre ASC;
        """)
        for row in cursor.fetchall():
            ws.append(list(row))
            
        for r_idx in range(4, ws.max_row + 1):
            ws.row_dimensions[r_idx].height = 20
            for c_idx in range(1, 5):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = normal_font
                cell.border = thin_border
                if c_idx == 3:
                    cell.alignment = Alignment(horizontal="center")
                    if cell.value == "ACTIVO":
                        cell.font = Font(name="Calibri", size=11, color="10B981", bold=True)
                    else:
                        cell.font = Font(name="Calibri", size=11, color="EF4444", bold=True)
                elif c_idx == 4:
                    cell.alignment = Alignment(horizontal="center")
        filename = "base_datos_personal.xlsx"
        
    elif tipo == "subnovedades":
        ws.title = "Catálogo Novedades"
        ws.merge_cells("A1:B1")
        ws["A1"] = "BIMEJ12 — CATÁLOGO DE SUBNOVEDADES"
        ws["A1"].font = title_font
        ws["A1"].fill = title_fill
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40
        
        headers = ["ID", "NOMBRE NOVEDAD"]
        ws.append([])
        ws.append(headers)
        ws.row_dimensions[3].height = 25
        
        for col_idx in range(1, 3):
            cell = ws.cell(row=3, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        cursor.execute("SELECT id, nombre FROM SUB_NOVEDADES ORDER BY nombre ASC;")
        for row in cursor.fetchall():
            ws.append(list(row))
            
        for r_idx in range(4, ws.max_row + 1):
            ws.row_dimensions[r_idx].height = 20
            for c_idx in range(1, 3):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = normal_font
                cell.border = thin_border
                if c_idx == 1:
                    cell.alignment = Alignment(horizontal="center")
        filename = "catalogo_subnovedades.xlsx"
        
    elif tipo == "consolidado_mensual" and mes:
        ws.title = f"Consolidado {mes}"
        dates = get_month_dates(mes)
        if not dates:
            raise HTTPException(status_code=400, detail="No hay reportes para el mes especificado.")
            
        placeholders = ",".join("%s" for _ in dates)
        cursor.execute(f"SELECT id, fecha FROM REPORTES WHERE fecha IN ({placeholders}) ORDER BY fecha ASC;", dates)
        reports_db = cursor.fetchall()
        report_ids = [r[0] for r in reports_db]
        report_dates = [r[1] for r in reports_db]
        
        num_cols = 2 + len(report_dates)
        col_letter = get_column_letter(num_cols)
        
        ws.merge_cells(f"A1:{col_letter}1")
        ws["A1"] = f"BIMEJ12 — CONSOLIDADO DIARIO DE PERSONAL — {mes.upper()}"
        ws["A1"].font = title_font
        ws["A1"].fill = title_fill
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40
        
        headers = ["CÉDULA", "INTEGRANTE"] + [f"Día {d.split('-')[2]}" for d in report_dates]
        ws.append([])
        ws.append(headers)
        ws.row_dimensions[3].height = 25
        
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        if report_ids:
            rep_placeholders = ",".join("%s" for _ in report_ids)
            cursor.execute(f"""
                SELECT p.cedula, p.nombre, rp.id_reporte, sn.nombre as subnovedad
                FROM REGISTRO_PERSONAL rp
                JOIN PERSONAL p ON rp.id_personal = p.id
                JOIN SUB_NOVEDADES sn ON rp.id_sub_novedad = sn.id
                WHERE rp.id_reporte IN ({rep_placeholders})
                ORDER BY p.nombre ASC;
            """, report_ids)
            
            person_map = {}
            for row in cursor.fetchall():
                key = (row[0], row[1])
                if key not in person_map:
                    person_map[key] = {}
                person_map[key][row[2]] = row[3]
                
            for (cedula, nombre), reports_dict in sorted(person_map.items(), key=lambda x: x[0][1]):
                row_data = [cedula, nombre]
                for r_id in report_ids:
                    row_data.append(reports_dict.get(r_id, "N/A"))
                ws.append(row_data)
                
        for r_idx in range(4, ws.max_row + 1):
            ws.row_dimensions[r_idx].height = 20
            for c_idx in range(1, num_cols + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = normal_font
                cell.border = thin_border
                if c_idx >= 3:
                    cell.alignment = Alignment(horizontal="center")
                    val = cell.value
                    if val in DISPONIBLE_STATUSES:
                        cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                        cell.font = Font(name="Calibri", size=9, color="065F46")
                    elif val == "N/A":
                        cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
                        cell.font = Font(name="Calibri", size=9, color="6B7280")
                    else:
                        cell.fill = PatternFill(start_color="FFE4E6", end_color="FFE4E6", fill_type="solid")
                        cell.font = Font(name="Calibri", size=9, color="991B1B")
        filename = f"consolidado_personal_{mes}.xlsx"
        
    elif tipo == "historial_novedades":
        ws.title = "Historial Novedades"
        ws.merge_cells("A1:G1")
        ws["A1"] = "BIMEJ12 — HISTORIAL COMPLETO DE NOVEDADES"
        ws["A1"].font = title_font
        ws["A1"].fill = title_fill
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40
        
        headers = ["CÉDULA", "APELLIDOS Y NOMBRES", "SUBNOVEDAD", "DESCRIPCIÓN", "DESDE", "HASTA", "FECHA REPORTE"]
        ws.append([])
        ws.append(headers)
        ws.row_dimensions[3].height = 25
        
        for col_idx in range(1, 8):
            cell = ws.cell(row=3, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        cursor.execute("""
            SELECT p.cedula, p.nombre, sn.nombre, rp.descripcion, rp.fecha_inicio, rp.fecha_final, r.fecha
            FROM REGISTRO_PERSONAL rp
            JOIN PERSONAL p ON rp.id_personal = p.id
            JOIN SUB_NOVEDADES sn ON rp.id_sub_novedad = sn.id
            JOIN REPORTES r ON rp.id_reporte = r.id
            ORDER BY r.fecha DESC, p.nombre ASC;
        """)
        for row in cursor.fetchall():
            ws.append(list(row))
            
        for r_idx in range(4, ws.max_row + 1):
            ws.row_dimensions[r_idx].height = 20
            for c_idx in range(1, 8):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = normal_font
                cell.border = thin_border
                if c_idx == 1:
                    cell.alignment = Alignment(horizontal="left")
                elif c_idx in (5, 6, 7):
                    cell.alignment = Alignment(horizontal="center")
        filename = "historial_completo_novedades.xlsx"
        
    else:
        raise HTTPException(status_code=400, detail="Parámetros inválidos")
        
    # Auto-adjust column widths
    if tipo != "consolidado_mensual":
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row == 1:
                    continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
    else:
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 28
        for col_idx in range(3, num_cols + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 8
        
    out_file = io.BytesIO()
    wb.save(out_file)
    out_file.seek(0)
    
    response = StreamingResponse(out_file, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response

@app.get("/api/exportar/pdf")
def exportar_pdf(
    tipo: str = Query(...),
    fecha: Optional[str] = Query(None),
    mes: Optional[str] = Query(None),
    cedula: Optional[int] = Query(None),
    db: sqlite3.Connection = Depends(get_db)
):
    pdf_buffer = io.BytesIO()
    
    # Select document layout based on export type
    doc_layout = landscape(letter) if tipo in ("dia", "mes", "consolidado_mensual", "historial_novedades") else letter
    doc = SimpleDocTemplate(
        pdf_buffer,
        pagesize=doc_layout,
        leftMargin=36,
        rightMargin=36,
        topMargin=36,
        bottomMargin=36
    )
    
    styles = getSampleStyleSheet()
    
    # Custom Styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=colors.HexColor('#0F172A'),
        alignment=1, # Center
        spaceAfter=15
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=11,
        textColor=colors.HexColor('#475569'),
        alignment=1, # Center
        spaceAfter=20
    )
    
    th_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.white,
        alignment=1 # Center
    )
    
    td_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        textColor=colors.HexColor('#1E293B'),
        alignment=0 # Left
    )
    
    story = []
    cursor = db.cursor()
    
    if tipo == "dia" and fecha:
        story.append(Paragraph(f"BIMEJ12 — REPORTE DIARIO DE PERSONAL", title_style))
        story.append(Paragraph(f"Fecha del Reporte: {fecha} | Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style))
        
        headers = [
            Paragraph("CÉDULA", th_style),
            Paragraph("APELLIDOS Y NOMBRES", th_style),
            Paragraph("SUBNOVEDAD", th_style),
            Paragraph("DESCRIPCIÓN", th_style),
            Paragraph("DESDE", th_style),
            Paragraph("HASTA", th_style)
        ]
        
        data = [headers]
        
        cursor.execute("SELECT id FROM REPORTES WHERE fecha = ?;", (fecha,))
        r_row = cursor.fetchone()
        if r_row:
            cursor.execute("""
                SELECT p.cedula, p.nombre, sn.nombre, rp.descripcion, rp.fecha_inicio, rp.fecha_final
                FROM REGISTRO_PERSONAL rp
                JOIN PERSONAL p ON rp.id_personal = p.id
                JOIN SUB_NOVEDADES sn ON rp.id_sub_novedad = sn.id
                WHERE rp.id_reporte = ?
                ORDER BY p.nombre ASC;
            """, (r_row[0],))
            
            for row in cursor.fetchall():
                data.append([
                    Paragraph(str(row[0]), td_style),
                    Paragraph(row[1], td_style),
                    Paragraph(row[2], td_style),
                    Paragraph(row[3] or "", td_style),
                    Paragraph(row[4] or "-", td_style),
                    Paragraph(row[5] or "-", td_style)
                ])
                
        col_widths = [75, 180, 110, 195, 80, 80]
        
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')])
        ]))
        story.append(t)
        filename = f"reporte_dia_{fecha}.pdf"
        
    elif tipo == "mes" and mes:
        story.append(Paragraph(f"BIMEJ12 — RESUMEN MENSUAL DE DISPONIBILIDAD", title_style))
        story.append(Paragraph(f"Mes: {mes.upper()} | Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style))
        
        headers = [
            Paragraph("FECHA", th_style),
            Paragraph("TOTAL PERSONAL", th_style),
            Paragraph("DISPONIBLES", th_style),
            Paragraph("EN NOVEDADES", th_style),
            Paragraph("DISPONIBILIDAD %", th_style)
        ]
        
        data = [headers]
        
        dates = get_month_dates(mes)
        for d in dates:
            cursor.execute("SELECT id FROM REPORTES WHERE fecha = ?;", (d,))
            r_id = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) FROM REGISTRO_PERSONAL WHERE id_reporte = ?;", (r_id,))
            total = cursor.fetchone()[0]
            
            placeholders = ",".join("?" for _ in DISPONIBLE_STATUSES)
            cursor.execute(f"""
                SELECT COUNT(*) FROM REGISTRO_PERSONAL 
                WHERE id_reporte = ? AND id_sub_novedad IN (
                    SELECT id FROM SUB_NOVEDADES WHERE nombre IN ({placeholders})
                );
            """, (r_id, *DISPONIBLE_STATUSES))
            disp = cursor.fetchone()[0]
            nov = total - disp
            pct = round((disp / total * 100), 1) if total > 0 else 0.0
            
            data.append([
                Paragraph(d, td_style),
                Paragraph(str(total), td_style),
                Paragraph(str(disp), td_style),
                Paragraph(str(nov), td_style),
                Paragraph(f"{pct}%", td_style)
            ])
            
        col_widths = [144, 144, 144, 144, 144] # 720 total
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')])
        ]))
        story.append(t)
        filename = f"reporte_mensual_{mes}.pdf"
        
    elif tipo == "personal" and cedula:
        cursor.execute("SELECT id, nombre, CASE WHEN fecha_retiro IS NULL THEN 'ACTIVO' ELSE 'RETIRADO' END as estado, fecha_retiro FROM PERSONAL WHERE cedula = ?;", (cedula,))
        p_row = cursor.fetchone()
        if not p_row:
            raise HTTPException(status_code=404, detail="Personal no encontrado")
            
        p_id, nombre, estado, fecha_retiro = p_row[0], p_row[1], p_row[2], p_row[3]
        
        story.append(Paragraph(f"HISTORIAL DE PERSONAL INDIVIDUAL", title_style))
        story.append(Paragraph(f"Integrante: {nombre} | Cédula: {cedula} | Estado: {estado} " + (f"| Fecha Retiro: {fecha_retiro}" if fecha_retiro else ""), subtitle_style))
        
        cursor.execute("SELECT COUNT(*) FROM REGISTRO_PERSONAL WHERE id_personal = ?;", (p_id,))
        total_dias = cursor.fetchone()[0]
        
        placeholders = ",".join("?" for _ in DISPONIBLE_STATUSES)
        cursor.execute(f"""
            SELECT COUNT(*) FROM REGISTRO_PERSONAL 
            WHERE id_personal = ? AND id_sub_novedad IN (
                SELECT id FROM SUB_NOVEDADES WHERE nombre IN ({placeholders})
            );
        """, (p_id, *DISPONIBLE_STATUSES))
        disp_dias = cursor.fetchone()[0]
        nov_dias = total_dias - disp_dias
        disp_pct = round((disp_dias / total_dias * 100), 1) if total_dias > 0 else 0.0
        
        stats_data = [
            [Paragraph("<b>Total Días Registrados:</b>", td_style), Paragraph(str(total_dias), td_style),
             Paragraph("<b>Días Disponibles:</b>", td_style), Paragraph(f"{disp_dias} ({disp_pct}%)", td_style)],
            [Paragraph("<b>Días en Novedades:</b>", td_style), Paragraph(f"{nov_dias} ({round(100.0 - disp_pct, 1)}%)", td_style),
             Paragraph("<b>Estado Actual:</b>", td_style), Paragraph(estado, td_style)]
        ]
        stats_table = Table(stats_data, colWidths=[135, 135, 135, 135])
        stats_table.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F8FAFC')),
            ('PADDING', (0,0), (-1,-1), 6)
        ]))
        story.append(stats_table)
        story.append(Spacer(1, 15))
        
        headers = [
            Paragraph("FECHA", th_style),
            Paragraph("SUBNOVEDAD", th_style),
            Paragraph("DESCRIPCIÓN", th_style),
            Paragraph("DESDE", th_style),
            Paragraph("HASTA", th_style)
        ]
        
        data = [headers]
        
        cursor.execute("""
            SELECT r.fecha, sn.nombre, rp.descripcion, rp.fecha_inicio, rp.fecha_final
            FROM REGISTRO_PERSONAL rp
            JOIN REPORTES r ON rp.id_reporte = r.id
            JOIN SUB_NOVEDADES sn ON rp.id_sub_novedad = sn.id
            WHERE rp.id_personal = ?
            ORDER BY r.fecha DESC;
        """, (p_id,))
        
        for row in cursor.fetchall():
            data.append([
                Paragraph(row[0], td_style),
                Paragraph(row[1], td_style),
                Paragraph(row[2] or "", td_style),
                Paragraph(row[3] or "-", td_style),
                Paragraph(row[4] or "-", td_style)
            ])
            
        col_widths = [75, 110, 195, 80, 80]
        
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')])
        ]))
        story.append(t)
        filename = f"historial_personal_{cedula}.pdf"
        
    elif tipo == "personal_db":
        doc_layout = letter
        doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=doc_layout,
            leftMargin=36,
            rightMargin=36,
            topMargin=36,
            bottomMargin=36
        )
        
        story.append(Paragraph("BIMEJ12 — BASE DE DATOS GENERAL DE PERSONAL", title_style))
        story.append(Paragraph(f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style))
        
        headers = [
            Paragraph("CÉDULA", th_style),
            Paragraph("APELLIDOS Y NOMBRES", th_style),
            Paragraph("ESTADO", th_style),
            Paragraph("FECHA RETIRO", th_style)
        ]
        data = [headers]
        
        cursor.execute("""
            SELECT cedula, nombre, CASE WHEN fecha_retiro IS NULL THEN 'ACTIVO' ELSE 'RETIRADO' END as estado, fecha_retiro
            FROM PERSONAL
            ORDER BY nombre ASC;
        """)
        
        active_style = ParagraphStyle('ActCell', parent=td_style, textColor=colors.HexColor('#10B981'), fontName='Helvetica-Bold')
        ret_style = ParagraphStyle('RetCell', parent=td_style, textColor=colors.HexColor('#EF4444'), fontName='Helvetica-Bold')
        
        for row in cursor.fetchall():
            est_text = row[2]
            est_p = Paragraph(est_text, active_style) if est_text == "ACTIVO" else Paragraph(est_text, ret_style)
            data.append([
                Paragraph(str(row[0]), td_style),
                Paragraph(row[1], td_style),
                est_p,
                Paragraph(row[3] or "-", td_style)
            ])
            
        col_widths = [90, 230, 110, 110]
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')])
        ]))
        story.append(t)
        filename = "base_datos_personal.pdf"

    elif tipo == "subnovedades":
        doc_layout = letter
        doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=doc_layout,
            leftMargin=54,
            rightMargin=54,
            topMargin=54,
            bottomMargin=54
        )
        
        story.append(Paragraph("BIMEJ12 — CATÁLOGO DE SUBNOVEDADES", title_style))
        story.append(Paragraph(f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style))
        
        headers = [
            Paragraph("ID", th_style),
            Paragraph("NOMBRE DE LA SUBNOVEDAD", th_style)
        ]
        data = [headers]
        
        cursor.execute("SELECT id, nombre FROM SUB_NOVEDADES ORDER BY nombre ASC;")
        for row in cursor.fetchall():
            data.append([
                Paragraph(str(row[0]), td_style),
                Paragraph(row[1], td_style)
            ])
            
        col_widths = [100, 404]
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')])
        ]))
        story.append(t)
        filename = "catalogo_subnovedades.pdf"
        
    elif tipo == "consolidado_mensual" and mes:
        doc_layout = landscape(letter)
        doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=doc_layout,
            leftMargin=36,
            rightMargin=36,
            topMargin=36,
            bottomMargin=36
        )
        
        story.append(Paragraph(f"BIMEJ12 — CONSOLIDADO DIARIO DE PERSONAL — {mes.upper()}", title_style))
        story.append(Paragraph(f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')} | D = Disponible, N = Novedad, - = N/A", subtitle_style))
        
        dates = get_month_dates(mes)
        if not dates:
            raise HTTPException(status_code=400, detail="No hay reportes para el mes especificado.")
            
        placeholders = ",".join("%s" for _ in dates)
        cursor.execute(f"SELECT id, fecha FROM REPORTES WHERE fecha IN ({placeholders}) ORDER BY fecha ASC;", dates)
        reports_db = cursor.fetchall()
        report_ids = [r[0] for r in reports_db]
        report_dates = [r[1] for r in reports_db]
        
        headers = [
            Paragraph("CÉDULA", th_style),
            Paragraph("INTEGRANTE", th_style)
        ] + [Paragraph(d.split('-')[2], th_style) for d in report_dates]
        
        data = [headers]
        
        disp_style = ParagraphStyle('DStyle', parent=td_style, textColor=colors.HexColor('#10B981'), fontName='Helvetica-Bold', alignment=1)
        nov_style = ParagraphStyle('NStyle', parent=td_style, textColor=colors.HexColor('#EF4444'), fontName='Helvetica-Bold', alignment=1)
        na_style = ParagraphStyle('NAStyle', parent=td_style, textColor=colors.HexColor('#6B7280'), alignment=1)
        
        if report_ids:
            rep_placeholders = ",".join("%s" for _ in report_ids)
            cursor.execute(f"""
                SELECT p.cedula, p.nombre, rp.id_reporte, sn.nombre as subnovedad
                FROM REGISTRO_PERSONAL rp
                JOIN PERSONAL p ON rp.id_personal = p.id
                JOIN SUB_NOVEDADES sn ON rp.id_sub_novedad = sn.id
                WHERE rp.id_reporte IN ({rep_placeholders})
                ORDER BY p.nombre ASC;
            """, report_ids)
            
            person_map = {}
            for row in cursor.fetchall():
                key = (row[0], row[1])
                if key not in person_map:
                    person_map[key] = {}
                person_map[key][row[2]] = row[3]
                
            for (cedula, nombre), reports_dict in sorted(person_map.items(), key=lambda x: x[0][1]):
                row_data = [
                    Paragraph(str(cedula), td_style),
                    Paragraph(nombre, td_style)
                ]
                for r_id in report_ids:
                    val = reports_dict.get(r_id, "N/A")
                    if val in DISPONIBLE_STATUSES:
                        row_data.append(Paragraph("D", disp_style))
                    elif val == "N/A":
                        row_data.append(Paragraph("-", na_style))
                    else:
                        row_data.append(Paragraph("N", nov_style))
                data.append(row_data)
                
        num_days_col = len(report_dates)
        day_col_width = 18 if num_days_col <= 31 else 15
        name_width = 720 - 60 - (num_days_col * day_col_width)
        col_widths = [60, name_width] + [day_col_width for _ in report_dates]
        
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
            ('ALIGN', (0,0), (1,-1), 'LEFT'),
            ('ALIGN', (2,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
            ('TOPPADDING', (0,0), (-1,-1), 3),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')])
        ]))
        story.append(t)
        filename = f"consolidado_personal_{mes}.pdf"
        
    elif tipo == "historial_novedades":
        doc_layout = landscape(letter)
        doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=doc_layout,
            leftMargin=36,
            rightMargin=36,
            topMargin=36,
            bottomMargin=36
        )
        
        story.append(Paragraph("BIMEJ12 — HISTORIAL COMPLETO DE NOVEDADES", title_style))
        story.append(Paragraph(f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style))
        
        headers = [
            Paragraph("CÉDULA", th_style),
            Paragraph("APELLIDOS Y NOMBRES", th_style),
            Paragraph("SUBNOVEDAD", th_style),
            Paragraph("DESCRIPCIÓN", th_style),
            Paragraph("DESDE", th_style),
            Paragraph("HASTA", th_style),
            Paragraph("REPORTE", th_style)
        ]
        
        data = [headers]
        
        cursor.execute("""
            SELECT p.cedula, p.nombre, sn.nombre, rp.descripcion, rp.fecha_inicio, rp.fecha_final, r.fecha
            FROM REGISTRO_PERSONAL rp
            JOIN PERSONAL p ON rp.id_personal = p.id
            JOIN SUB_NOVEDADES sn ON rp.id_sub_novedad = sn.id
            JOIN REPORTES r ON rp.id_reporte = r.id
            ORDER BY r.fecha DESC, p.nombre ASC;
        """)
        
        for row in cursor.fetchall():
            data.append([
                Paragraph(str(row[0]), td_style),
                Paragraph(row[1], td_style),
                Paragraph(row[2], td_style),
                Paragraph(row[3] or "", td_style),
                Paragraph(row[4] or "-", td_style),
                Paragraph(row[5] or "-", td_style),
                Paragraph(row[6], td_style)
            ])
            
        col_widths = [70, 150, 100, 160, 80, 80, 80]
        
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')])
        ]))
        story.append(t)
        filename = "historial_completo_novedades.pdf"
        
    else:
        raise HTTPException(status_code=400, detail="Parámetros inválidos")
        
    doc.build(story)
    pdf_buffer.seek(0)
    
    response = StreamingResponse(pdf_buffer, media_type="application/pdf")
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response

# Standard start helper if run directly
if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="127.0.0.1", port=8000, reload=True)
